"""
Тестовый скрипт для проверки генерации Excel-отчета со статистикой
"""
import os
import sys
from datetime import datetime, timedelta
from dotenv import load_dotenv
from db import get_session
from model import Order, User, City, Stat
from services.commission_service import get_master_pct, get_category_for_type

# Загружаем переменные окружения
load_dotenv('local.env')

def generate_test_excel():
    """Генерирует Excel-отчет на основе тестовых данных"""
    session = get_session()
    
    try:
        # Получаем все заявки со статусом completed (или done_pending_sum для теста)
        orders = session.query(Order).filter(
            Order.status.in_(["completed", "done_pending_sum"])
        ).order_by(Order.created_at.desc()).all()
        
        if not orders:
            print("❌ Нет заявок для генерации отчета.")
            print("💡 Сначала запустите test_stats_generator.py для создания тестовых данных")
            return None
        
        print(f"📊 Найдено {len(orders)} заявок для отчета")
        
        # Генерируем Excel используя ту же логику, что и в vk_bot.py
        import tempfile
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика"
        
        # Заголовки
        headers = [
            "Номер", "Город", "Дата", "Тип", "Категория",
            "Сумма", "ЗПЧ", "СД", "Чистый чек", "% мастера",
            "Доля мастера", "Доля компании", "Статус", "Мастер VK"
        ]
        ws.append(headers)
        
        # Форматирование заголовков
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные
        print("📝 Заполнение данных...")
        for o in orders:
            city_name = o.city_rel.name if getattr(o, 'city_rel', None) else "-"
            order_sum = float(getattr(o, 'sum_amount', 0) or 0)
            
            # Безопасное получение sd_price и zpch_sum
            try:
                sd_price = float(getattr(o, 'sd_price', 0) or 0)
            except (AttributeError, ValueError, TypeError):
                sd_price = 0.0
            
            try:
                zpch = float(getattr(o, 'zpch_sum', 0) or 0)
            except (AttributeError, ValueError, TypeError):
                zpch = 0.0
            
            net = max(order_sum - sd_price - zpch, 0)
            
            try:
                pct = get_master_pct(o.equip_type, net)
            except Exception:
                pct = 40.0
            
            master_share = net * (pct / 100.0)
            company_share = max(net - master_share, 0)
            cat = get_category_for_type(o.equip_type)
            
            date_val = o.created_at.date() if getattr(o, 'created_at', None) else None
            
            ws.append([
                int(o.order_number) if o.order_number else "-",
                str(city_name) if city_name else "-",
                date_val if date_val else "-",
                str(o.equip_type) if o.equip_type else "-",
                str(cat) if cat else "-",
                round(order_sum, 2),
                round(zpch, 2),
                round(sd_price, 2),
                round(net, 2),
                round(pct, 2),
                round(master_share, 2),
                round(company_share, 2),
                str(o.status) if o.status else "-",
                int(o.assigned_to) if o.assigned_to else "-"
            ])
        
        # Автоподбор ширины колонок
        print("📏 Настройка ширины колонок...")
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Создаем лист со статистикой по категориям и диапазонам
        ws_stats = wb.create_sheet("Статистика по категориям")
        
        # Загружаем настройки комиссий
        from services.commission_service import load_settings
        settings = load_settings()
        
        # Заголовки для статистики
        ws_stats.append(["Категория", "Диапазон", "Процент", "Кол-во заявок", "Сумма чистого чека", "Доля мастера", "Доля компании"])
        
        # Форматирование заголовков статистики
        for cell in ws_stats[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Группируем заявки по категориям и диапазонам
        stats_by_category = {}
        
        for o in orders:
            try:
                order_sum = float(getattr(o, 'sum_amount', 0) or 0)
                sd_price = float(getattr(o, 'sd_price', 0) or 0)
                zpch = float(getattr(o, 'zpch_sum', 0) or 0)
                net = max(order_sum - sd_price - zpch, 0)
                
                cat = get_category_for_type(o.equip_type)
                pct = get_master_pct(o.equip_type, net)
                
                # Определяем диапазон
                cat_settings = settings.get(cat, {})
                tiers = cat_settings.get("tiers", [])
                range_str = "Не определен"
                
                for lo, hi, tier_pct in tiers:
                    if net >= (lo or 0) and (hi is None or net <= hi):
                        if hi is None:
                            range_str = f"от {lo}+"
                        else:
                            range_str = f"{lo} - {hi}"
                        break
                
                # Группируем
                key = (cat, range_str, pct)
                if key not in stats_by_category:
                    stats_by_category[key] = {
                        "count": 0,
                        "net_sum": 0.0,
                        "master_sum": 0.0,
                        "company_sum": 0.0
                    }
                
                master_share = net * (pct / 100.0)
                company_share = max(net - master_share, 0)
                
                stats_by_category[key]["count"] += 1
                stats_by_category[key]["net_sum"] += net
                stats_by_category[key]["master_sum"] += master_share
                stats_by_category[key]["company_sum"] += company_share
                
            except Exception as e:
                print(f"Ошибка при обработке заявки {o.id}: {e}")
        
        # Записываем статистику
        for (cat, range_str, pct), data in sorted(stats_by_category.items()):
            cat_title = settings.get(cat, {}).get("title", cat)
            ws_stats.append([
                cat_title,
                range_str,
                f"{pct}%",
                data["count"],
                round(data["net_sum"], 2),
                round(data["master_sum"], 2),
                round(data["company_sum"], 2)
            ])
        
        # Автоподбор ширины для листа статистики
        for col_idx, col in enumerate(ws_stats.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_stats.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение файла
        filename = f"test_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        
        wb.save(filepath)
        print(f"✅ Excel-файл создан: {filepath}")
        print(f"📊 Всего строк в основном листе: {len(orders) + 1} (включая заголовок)")
        print(f"📈 Статистика по категориям: {len(stats_by_category)} записей")
        print(f"📁 Полный путь: {os.path.abspath(filepath)}")
        
        # Выводим статистику в консоль
        print("\n📈 Статистика по категориям и диапазонам:")
        for (cat, range_str, pct), data in sorted(stats_by_category.items()):
            cat_title = settings.get(cat, {}).get("title", cat)
            print(f"  {cat_title} | {range_str} ({pct}%):")
            print(f"    Заявок: {data['count']}, Чистый чек: {data['net_sum']:.2f} руб.")
            print(f"    Доля мастера: {data['master_sum']:.2f} руб., Доля компании: {data['company_sum']:.2f} руб.")
        
        # Пытаемся автоматически открыть файл
        try:
            import platform
            import subprocess
            system = platform.system()
            if system == "Windows":
                os.startfile(filepath)
                print(f"\n🖥️ Файл открыт в Excel")
            elif system == "Darwin":  # macOS
                subprocess.run(["open", filepath])
                print(f"\n🖥️ Файл открыт")
            elif system == "Linux":
                subprocess.run(["xdg-open", filepath])
                print(f"\n🖥️ Файл открыт")
        except Exception as e:
            print(f"\n💡 Файл не открылся автоматически: {e}")
            print(f"   Откройте файл вручную: {filepath}")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Ошибка при генерации Excel: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        session.close()


def generate_from_stats():
    """Генерирует Excel на основе записей Stat (если есть)"""
    session = get_session()
    
    try:
        stats = session.query(Stat).all()
        
        if not stats:
            print("❌ Нет записей в таблице Stat.")
            print("💡 Записи в Stat создаются при приемке кассы")
            return None
        
        print(f"📊 Найдено {len(stats)} записей в Stat")
        
        # Создаем Excel
        import tempfile
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика из Stat"
        
        headers = [
            "ID", "Order ID", "Тип", "Сумма", "Отказ", "Мастер VK", "Дата записи"
        ]
        ws.append(headers)
        
        # Форматирование заголовков
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for s in stats:
            ws.append([
                s.id,
                s.order_id,
                s.equip_type or "-",
                round(s.sum or 0, 2),
                "Да" if s.refused else "Нет",
                s.master_tg or "-",
                s.recorded_at.date() if s.recorded_at else "-"
            ])
        
        # Автоподбор ширины
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        filename = f"test_stats_from_stat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        
        wb.save(filepath)
        print(f"✅ Excel-файл создан: {filepath}")
        print(f"📁 Полный путь: {os.path.abspath(filepath)}")
        
        # Пытаемся автоматически открыть файл
        try:
            import platform
            import subprocess
            system = platform.system()
            if system == "Windows":
                os.startfile(filepath)
                print(f"\n🖥️ Файл открыт в Excel")
            elif system == "Darwin":  # macOS
                subprocess.run(["open", filepath])
                print(f"\n🖥️ Файл открыт")
            elif system == "Linux":
                subprocess.run(["xdg-open", filepath])
                print(f"\n🖥️ Файл открыт")
        except Exception as e:
            print(f"\n💡 Файл не открылся автоматически: {e}")
            print(f"   Откройте файл вручную: {filepath}")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        session.close()


if __name__ == "__main__":
    print("=" * 60)
    print("📊 Тест генерации Excel-отчета")
    print("=" * 60)
    print()
    
    if len(sys.argv) > 1 and sys.argv[1] == "stat":
        # Генерируем из таблицы Stat
        print("📋 Генерация из таблицы Stat...")
        generate_from_stats()
    else:
        # Генерируем из заявок (как в боте)
        print("📋 Генерация из заявок (как в боте)...")
        filepath = generate_test_excel()
        
        if filepath:
            print(f"\n💡 Для просмотра откройте файл: {filepath}")
            print(f"💡 Или используйте: python test_excel_generation.py stat - для генерации из Stat")

