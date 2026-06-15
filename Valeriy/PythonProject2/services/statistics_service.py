from __future__ import annotations

import os
import tempfile
from collections import OrderedDict
from datetime import datetime
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.commission_service import get_category_for_type, load_settings


CATEGORY_ORDER: List[str] = ["appliance", "phones", "pc", "other"]
HEADERS: List[str] = [
    "Город",
    "Всего",
    "Отказ",
    "Гарантия",
    "Оборот",
    "Средний чек",
    "KPI директора",
]


def _safe_city_name(order) -> str:
    city_rel = getattr(order, "city_rel", None)
    if city_rel and getattr(city_rel, "name", None):
        return city_rel.name
    return getattr(order, "city_name", "Без города")


def _calculate_net(order) -> float:
    sum_amount = float(getattr(order, "sum_amount", 0) or 0)
    sd_price = float(getattr(order, "sd_price", 0) or 0)
    zpch_sum = float(getattr(order, "zpch_sum", 0) or 0)
    net = sum_amount - sd_price - zpch_sum
    return round(max(net, 0.0), 2)


def collect_city_stats(orders: Iterable) -> Dict[str, Dict]:
    settings = load_settings()
    category_titles = {
        code: settings.get(code, {}).get("title", code) for code in CATEGORY_ORDER
    }

    stats: Dict[str, Dict] = OrderedDict()

    for order in orders:
        if getattr(order, "status", "") != "completed":
            continue

        city_name = _safe_city_name(order)
        city_stat = stats.setdefault(
            city_name,
            {
                "total": 0,
                "refused": 0,
                "warranty": 0,
                "turnover": 0.0,
                "categories": {
                    code: {
                        "title": category_titles.get(code, code),
                        "total": 0,
                        "refused": 0,
                        "warranty": 0,
                        "turnover": 0.0,
                    }
                    for code in CATEGORY_ORDER
                },
            },
        )

        net_amount = _calculate_net(order)
        is_warranty = bool(getattr(order, "is_warranty", False))
        sum_amount = float(getattr(order, "sum_amount", 0) or 0)
        is_refused = sum_amount <= 0 and not is_warranty

        if not is_warranty:
            city_stat["total"] += 1
        else:
            city_stat["warranty"] += 1

        if is_refused:
            city_stat["refused"] += 1

        city_stat["turnover"] += net_amount

        category_code = get_category_for_type(getattr(order, "equip_type", "other"))
        category_stat = city_stat["categories"][category_code]

        if not is_warranty:
            category_stat["total"] += 1
        else:
            category_stat["warranty"] += 1

        if is_refused:
            category_stat["refused"] += 1

        category_stat["turnover"] += net_amount

    for city_stat in stats.values():
        total = city_stat["total"]
        city_stat["avg_check"] = round(city_stat["turnover"] / total, 2) if total else 0.0
        city_stat["kpi"] = round(city_stat["turnover"] * 0.10, 2)

        for category_stat in city_stat["categories"].values():
            total_cat = category_stat["total"]
            category_stat["avg_check"] = (
                round(category_stat["turnover"] / total_cat, 2) if total_cat else 0.0
            )

    return stats


def _build_workbook(stats: Dict[str, Dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика по городам"

    ws.append(HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6AA84F")
    city_fill = PatternFill("solid", fgColor="B7E1CD")
    kpi_fill = PatternFill("solid", fgColor="6D9EEB")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for city_name, city_stat in stats.items():
        city_row = [
            city_name,
            city_stat["total"],
            city_stat["refused"],
            city_stat["warranty"],
            round(city_stat["turnover"], 2),
            round(city_stat["avg_check"], 2),
            round(city_stat["kpi"], 2),
        ]
        ws.append(city_row)
        row_idx = ws.max_row
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.fill = city_fill
            cell.font = Font(bold=True)
            if col_idx == len(HEADERS):
                cell.fill = kpi_fill
                cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for code in CATEGORY_ORDER:
            category_stat = city_stat["categories"].get(code)
            if not category_stat:
                continue
            ws.append(
                [
                    category_stat["title"],
                    category_stat["total"],
                    category_stat["refused"],
                    category_stat["warranty"],
                    round(category_stat["turnover"], 2),
                    round(category_stat["avg_check"], 2),
                    "",
                ]
            )
            row_idx = ws.max_row
            ws[row_idx][0].alignment = Alignment(horizontal="left")
            for cell in ws[row_idx][1:6]:
                cell.alignment = Alignment(horizontal="center")

        ws.append([])

    for col_idx, column in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    return wb


def generate_city_stats_excel_from_stats(
    stats: Dict[str, Dict], output_path: str | None = None
) -> str:
    if not stats:
        raise ValueError("Нет данных для формирования отчёта")

    wb = _build_workbook(stats)
    if output_path is None:
        tmp = tempfile.NamedTemporaryFile(prefix="city_stats_", suffix=".xlsx", delete=False)
        output_path = tmp.name
        tmp.close()

    wb.save(output_path)
    return output_path


def generate_city_stats_excel(orders: Iterable, output_path: str | None = None) -> str:
    stats = collect_city_stats(orders)
    return generate_city_stats_excel_from_stats(stats, output_path=output_path)
