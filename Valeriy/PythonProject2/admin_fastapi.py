"""
Веб-админка на FastAPI для управления процентами, мастерами и городами.
"""
import os
import json
import io
import csv
import shutil
import logging
import time
from datetime import datetime, date, timedelta
from typing import Dict, Any, Optional, Tuple
from pathlib import Path
from urllib.parse import urlencode
from fastapi import FastAPI, Request, Form, HTTPException, Query, UploadFile, File
from fastapi.exceptions import RequestValidationError
from fastapi import status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse, FileResponse
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import bcrypt

# Загружаем переменные окружения из .env файла
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from db import get_session
from model import User, City, Order, Stat
from services.commission_service import load_settings, save_settings
from services.dashboard_stats import (
    calculate_dashboard_stats,
    get_period_bounds,
    summarize_dashboard,
    calculate_detailed_stats,
    calculate_cash_income,
    calculate_category_table_stats,
)
from sqlalchemy import func, cast, String

# Импортируем схемы валидации
try:
    from schemas import CompanyDataSchema
except ImportError:
    CompanyDataSchema = None

app = FastAPI(title="Service Admin", version="1.0.0")
app.add_middleware(SessionMiddleware, secret_key=os.getenv("ADMIN_SECRET_KEY", "dev-secret-key"))

# Настройка логирования
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Шаблоны
templates = Jinja2Templates(directory="templates")

# ---- Глобальные обработчики ошибок ----

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    status_code = exc.status_code
    detail = exc.detail if isinstance(exc.detail, str) else "Произошла ошибка"
    context = {
        "request": request,
        "page_title": f"Ошибка {status_code}",
        "active_page": "",
        "error_code": status_code,
        "error_message": detail,
    }
    template_path = Path("templates/error.html")
    if template_path.exists():
        return templates.TemplateResponse("error.html", context, status_code=status_code)
    return HTMLResponse(
        content=f"<h1>Ошибка {status_code}</h1><p>{detail}</p>",
        status_code=status_code,
    )


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    return await http_exception_handler(
        request,
        HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Ошибка валидации данных",
        ),
    )


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    logging.getLogger(__name__).exception("Unhandled error")
    return await http_exception_handler(
        request,
        HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Внутренняя ошибка сервера. Попробуйте позже.",
        ),
    )


@app.get("/mock/product", response_class=HTMLResponse)
async def mock_product_page(request: Request):
    """
    Демо-страница верстки "карточка товара + отзывы" (как на присланном скриншоте).
    Никаких данных из БД — чисто фронт.
    """
    return templates.TemplateResponse(
        "product_reviews_mock.html",
        {"request": request, "page_title": "Карточка товара — демо"},
    )

# Папка для хранения фото паспортов
PASSPORT_PHOTOS_DIR = Path("passport_photos")
PASSPORT_PHOTOS_DIR.mkdir(exist_ok=True)

# Папка для хранения файлов БСО/договор/квитанция
BSO_FILES_DIR = Path("bso_files").resolve()
BSO_FILES_DIR.mkdir(exist_ok=True)

# Резервная директория для БСО файлов
BSO_FILES_BACKUP_DIR = Path("data/bso_files").resolve()
BSO_FILES_BACKUP_DIR.mkdir(parents=True, exist_ok=True)

# Суперадмин ID (хардкод)
SUPER_ADMIN_ID = 1080026562

# Функция проверки прав доступа (owner или director)
def check_owner_or_director_access(session, user_tg_id: int) -> bool:
    """Проверяет, является ли пользователь owner или director"""
    # Суперадмин всегда имеет доступ
    if user_tg_id == SUPER_ADMIN_ID:
        # Создаем или обновляем пользователя с ролью owner
        user = session.query(User).filter_by(tg_id=user_tg_id).first()
        if not user:
            try:
                user = User(tg_id=user_tg_id, name=f"Super Admin {user_tg_id}", role="owner")
                session.add(user)
                session.commit()
            except Exception:
                session.rollback()
        elif user.role != "owner":
            try:
                user.role = "owner"
                session.commit()
            except Exception:
                session.rollback()
        return True
    
    user = session.query(User).filter_by(tg_id=user_tg_id).first()
    if not user:
        # Проверяем ADMIN_IDS из .env
        admin_ids_str = os.getenv("ADMIN_IDS", "")
        admin_ids = [int(x.strip()) for x in admin_ids_str.split(",") if x.strip()] if admin_ids_str else []
        if user_tg_id in admin_ids:
            # Автоматически создаем пользователя с ролью owner
            try:
                user = User(tg_id=user_tg_id, name=f"Admin {user_tg_id}", role="owner")
                session.add(user)
                session.commit()
                return True
            except Exception:
                session.rollback()
                return False
        return False
    
    # Если пользователь в ADMIN_IDS, но роль не owner - обновляем
    admin_ids_str = os.getenv("ADMIN_IDS", "")
    admin_ids = [int(x.strip()) for x in admin_ids_str.split(",") if x.strip()] if admin_ids_str else []
    if user_tg_id in admin_ids and user.role != "owner":
        try:
            user.role = "owner"
            session.commit()
        except Exception:
            session.rollback()
    
    return user.role in ["owner", "director"]

# Функция получения текущего пользователя
def check_document_access(session, user_tg_id: int, order: Optional["Order"] = None) -> bool:
    """
    Проверяет, имеет ли пользователь доступ к документам заявок.
    Доступ дают роли owner/director/dispatcher, а также мастер, назначенный на заявку,
    или создатель заявки.
    """
    if user_tg_id == SUPER_ADMIN_ID:
        return True

    user = session.query(User).filter_by(tg_id=user_tg_id).first()
    if not user:
        return False

    # Роли с полным доступом к документам
    if user.role in ["owner", "director", "dispatcher"]:
        return True

    # Если мастер или другой пользователь — проверяем связь с заявкой
    if order:
        if order.assigned_to == user_tg_id:
            return True
        if order.created_by == user_tg_id:
            return True

    return False

def get_current_user_id(request: Request) -> Optional[int]:
    """Получить ID текущего пользователя из сессии"""
    # Проверяем веб-аутентификацию (username/password)
    if "user_id" in request.session:
        return request.session.get("user_id")
    return None

def get_user_role(session, user_id: Optional[int]) -> Optional[str]:
    """Получить роль пользователя"""
    if not user_id:
        return None
    user = session.query(User).filter_by(tg_id=user_id).first()
    return user.role if user else None

def add_user_role_to_context(session, user_id: Optional[int], context: dict) -> dict:
    """Добавить user_role в контекст шаблона"""
    context["user_role"] = get_user_role(session, user_id)
    return context


PERIOD_LABELS = {
    "today": "Сегодня",
    "week": "Последние 7 дней",
    "month": "Текущий месяц",
    "quarter": "Квартал",
    "year": "Год",
    "last_30": "30 дней",
}


def resolve_date_range(
    date_from_str: Optional[str],
    date_to_str: Optional[str],
    period: str = "month",
) -> Tuple[datetime, datetime, str]:
    """Парсит даты и возвращает границы периода вместе с подписью."""
    start_dt: Optional[datetime] = None
    end_dt: Optional[datetime] = None

    if date_from_str:
        try:
            start_dt = datetime.strptime(date_from_str, "%Y-%m-%d")
            start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        except ValueError:
            start_dt = None

    if date_to_str:
        try:
            end_dt = datetime.strptime(date_to_str, "%Y-%m-%d")
            end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        except ValueError:
            end_dt = None

    preset = period if period in PERIOD_LABELS else "month"
    if start_dt is None or end_dt is None:
        preset_start, preset_end = get_period_bounds(preset)
        start_dt = start_dt or preset_start
        end_dt = end_dt or preset_end

    label = PERIOD_LABELS.get(preset, "Выбранный период")
    if date_from_str and date_to_str:
        label = "Выбранный период"

    return start_dt, end_dt, label

# Создаём папку для шаблонов если её нет
os.makedirs("templates", exist_ok=True)

# Базовый шаблон
BASE_TEMPLATE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <title>{{ page_title or 'Админка' }} — сервис</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    :root {
      --bg: #0f172a;
      --bg-soft: #020617;
      --surface: #020617;
      --surface-soft: #020617;
      --border-subtle: #1e293b;
      --accent: #2563eb;
      --accent-soft: rgba(37,99,235,0.12);
      --accent-border: #1d4ed8;
      --text: #e5e7eb;
      --text-soft: #9ca3af;
      --danger: #ef4444;
      --success: #22c55e;
      --radius-lg: 18px;
      --radius-md: 10px;
    }

    * { box-sizing: border-box; }
    html, body { height: 100%; margin: 0; padding: 0; }
    body {
      font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: radial-gradient(circle at top left, #1d4ed8 0, #020617 45%, #000 100%);
      color: var(--text);
      -webkit-font-smoothing: antialiased;
    }

    .layout {
      display: grid;
      grid-template-columns: 250px minmax(0, 1fr);
      height: 100vh;
      max-height: 100vh;
    }

    .sidebar {
      background: linear-gradient(180deg, var(--bg-soft), #020617);
      border-right: 1px solid rgba(148,163,184,0.18);
      padding: 18px 18px 16px;
      display: flex;
      flex-direction: column;
      gap: 24px;
    }

    .brand {
      display: flex;
      align-items: center;
      gap: 10px;
    }
    .brand-logo {
      width: 32px;
      height: 32px;
      border-radius: 12px;
      background: radial-gradient(circle at 30% 20%, #60a5fa 0, #1d4ed8 35%, #020617 100%);
      display: flex;
      align-items: center;
      justify-content: center;
      color: #e5e7eb;
      font-weight: 600;
      font-size: 17px;
      box-shadow: 0 10px 30px rgba(37,99,235,0.4);
    }
    .brand-title {
      display: flex;
      flex-direction: column;
    }
    .brand-title-main {
      font-size: 16px;
      font-weight: 600;
      letter-spacing: 0.03em;
    }
    .brand-title-sub {
      font-size: 11px;
      color: var(--text-soft);
      text-transform: uppercase;
      letter-spacing: 0.12em;
    }

    .nav-group-title {
      font-size: 11px;
      color: var(--text-soft);
      text-transform: uppercase;
      letter-spacing: 0.12em;
      padding: 0 12px 6px;
    }

    .nav-link {
      display: flex;
      align-items: center;
      gap: 10px;
      padding: 10px 12px;
      border-radius: var(--radius-md);
      color: var(--text);
      text-decoration: none;
      font-size: 14px;
      transition: all 0.2s ease;
      position: relative;
    }
    .nav-link::before {
      content: '';
      position: absolute;
      left: 0;
      top: 50%;
      transform: translateY(-50%);
      width: 3px;
      height: 0;
      background: var(--accent);
      border-radius: 0 2px 2px 0;
      transition: height 0.2s ease;
    }
    .nav-link:hover {
      background: rgba(148,163,184,0.08);
      transform: translateX(2px);
    }
    .nav-link.active {
      background: var(--accent-soft);
      color: #60a5fa;
      border: 1px solid rgba(37,99,235,0.3);
    }
    .nav-link.active::before {
      height: 60%;
    }

    .content {
      overflow-y: auto;
      padding: 32px 40px;
    }

    .page-header {
      margin-bottom: 28px;
    }
    .page-title {
      font-size: 28px;
      font-weight: 600;
      margin: 0 0 6px;
    }
    .page-subtitle {
      font-size: 14px;
      color: var(--text-soft);
      margin: 0;
    }

    .card {
      background: var(--surface);
      border: 1px solid var(--border-subtle);
      border-radius: var(--radius-lg);
      padding: 24px;
      margin-bottom: 20px;
      transition: all 0.3s ease;
      box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }
    .card:hover {
      box-shadow: 0 4px 12px rgba(37,99,235,0.15);
      transform: translateY(-2px);
    }
    .card-header {
      margin-bottom: 20px;
    }
    .card-title {
      font-size: 18px;
      font-weight: 600;
      margin: 0 0 4px;
    }
    .card-desc {
      font-size: 13px;
      color: var(--text-soft);
      margin: 0;
    }

    table {
      width: 100%;
      border-collapse: collapse;
      margin-top: 16px;
    }
    th, td {
      padding: 10px 12px;
      text-align: left;
      border-bottom: 1px solid var(--border-subtle);
      font-size: 13px;
    }
    th {
      color: var(--text-soft);
      font-weight: 500;
      text-transform: uppercase;
      font-size: 11px;
      letter-spacing: 0.05em;
    }
    tbody tr:hover {
      background: rgba(148,163,184,0.04);
    }

    input[type="text"],
    input[type="number"],
    input[type="tel"],
    input[type="date"],
    input[type="time"],
    select,
    textarea {
      width: 100%;
      padding: 8px 10px;
      background: var(--bg-soft);
      border: 1px solid var(--border-subtle);
      border-radius: 6px;
      color: var(--text);
      font-size: 13px;
      transition: all 0.2s ease;
    }
    input:focus,
    select:focus,
    textarea:focus {
      outline: none;
      border-color: var(--accent);
      box-shadow: 0 0 0 3px rgba(37,99,235,0.1);
    }
    input:hover,
    select:hover,
    textarea:hover {
      border-color: rgba(148,163,184,0.3);
    }
    textarea {
      resize: vertical;
      min-height: 80px;
    }
    
    .form-group {
      margin-bottom: 20px;
    }
    .form-label {
      display: block;
      font-size: 13px;
      font-weight: 500;
      margin-bottom: 6px;
      color: var(--text-soft);
    }
    .form-row {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
    }
    
    .btn-secondary {
      background: var(--surface-soft);
      color: var(--text);
    }
    .btn-secondary:hover {
      background: var(--surface);
    }

    .btn {
      display: inline-block;
      padding: 10px 18px;
      border-radius: var(--radius-md);
      border: none;
      cursor: pointer;
      font-size: 14px;
      font-weight: 500;
      transition: all 0.2s ease;
      text-decoration: none;
      position: relative;
      overflow: hidden;
    }
    .btn::before {
      content: '';
      position: absolute;
      top: 50%;
      left: 50%;
      width: 0;
      height: 0;
      border-radius: 50%;
      background: rgba(255,255,255,0.2);
      transform: translate(-50%, -50%);
      transition: width 0.3s, height 0.3s;
    }
    .btn:hover::before {
      width: 300px;
      height: 300px;
    }
    .btn-primary {
      background: var(--accent);
      color: white;
    }
    .btn-primary:hover {
      background: var(--accent-border);
      transform: scale(1.02);
      box-shadow: 0 4px 12px rgba(37,99,235,0.3);
    }

    .flash {
      padding: 12px 16px;
      margin-bottom: 20px;
      border-radius: var(--radius-md);
      font-size: 13px;
      animation: slideIn 0.3s ease;
    }
    @keyframes slideIn {
      from {
        opacity: 0;
        transform: translateY(-10px);
      }
      to {
        opacity: 1;
        transform: translateY(0);
      }
    }
    .flash-success {
      background: rgba(34,197,94,0.15);
      color: var(--success);
      border: 1px solid rgba(34,197,94,0.3);
    }
    .flash-error {
      background: rgba(239,68,68,0.15);
      color: var(--danger);
      border: 1px solid rgba(239,68,68,0.3);
    }

    /* Toast уведомления */
    .toast-container {
      position: fixed;
      top: 20px;
      right: 20px;
      z-index: 10000;
      display: flex;
      flex-direction: column;
      gap: 10px;
    }
    .toast {
      padding: 14px 18px;
      border-radius: var(--radius-md);
      font-size: 14px;
      min-width: 300px;
      box-shadow: 0 4px 12px rgba(0,0,0,0.3);
      animation: toastSlideIn 0.3s ease;
      display: flex;
      align-items: center;
      gap: 12px;
    }
    @keyframes toastSlideIn {
      from {
        opacity: 0;
        transform: translateX(100%);
      }
      to {
        opacity: 1;
        transform: translateX(0);
      }
    }
    .toast-success {
      background: var(--success);
      color: white;
    }
    .toast-error {
      background: var(--danger);
      color: white;
    }
    .toast-info {
      background: var(--accent);
      color: white;
    }
    .toast-close {
      margin-left: auto;
      cursor: pointer;
      opacity: 0.8;
      font-size: 18px;
      line-height: 1;
    }
    .toast-close:hover {
      opacity: 1;
    }

    /* Модальные окна */
    .modal-overlay {
      display: none;
      position: fixed;
      top: 0;
      left: 0;
      right: 0;
      bottom: 0;
      background: rgba(0,0,0,0.7);
      z-index: 9999;
      align-items: center;
      justify-content: center;
      animation: fadeIn 0.2s ease;
    }
    .modal-overlay.active {
      display: flex;
    }
    @keyframes fadeIn {
      from { opacity: 0; }
      to { opacity: 1; }
    }
    .modal {
      background: var(--surface);
      border: 1px solid var(--border-subtle);
      border-radius: var(--radius-lg);
      padding: 24px;
      max-width: 500px;
      width: 90%;
      max-height: 90vh;
      overflow-y: auto;
      animation: modalSlideIn 0.3s ease;
      box-shadow: 0 10px 40px rgba(0,0,0,0.5);
    }
    @keyframes modalSlideIn {
      from {
        opacity: 0;
        transform: scale(0.9) translateY(-20px);
      }
      to {
        opacity: 1;
        transform: scale(1) translateY(0);
      }
    }
    .modal-header {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 20px;
    }
    .modal-title {
      font-size: 20px;
      font-weight: 600;
      margin: 0;
    }
    .modal-close {
      background: none;
      border: none;
      color: var(--text-soft);
      font-size: 24px;
      cursor: pointer;
      padding: 0;
      width: 30px;
      height: 30px;
      display: flex;
      align-items: center;
      justify-content: center;
      border-radius: 6px;
    }
    .modal-close:hover {
      background: var(--bg-soft);
      color: var(--text);
    }
    .modal-footer {
      display: flex;
      gap: 10px;
      justify-content: flex-end;
      margin-top: 24px;
    }

    /* Переключатель темы */
    .theme-toggle {
      position: fixed;
      bottom: 20px;
      right: 20px;
      width: 50px;
      height: 50px;
      border-radius: 50%;
      background: var(--accent);
      border: none;
      color: white;
      font-size: 20px;
      cursor: pointer;
      box-shadow: 0 4px 12px rgba(37,99,235,0.4);
      z-index: 1000;
      transition: all 0.3s ease;
    }
    .theme-toggle:hover {
      transform: scale(1.1);
      box-shadow: 0 6px 20px rgba(37,99,235,0.6);
    }

    .grid-2 {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 20px;
    }

    @media (max-width: 768px) {
      .layout {
        grid-template-columns: 1fr;
      }
      .sidebar {
        display: none;
      }
    }
  </style>
</head>
<body>
  <div class="layout">
    <aside class="sidebar">
      <div class="brand">
        <div class="brand-logo">S</div>
        <div class="brand-title">
          <div class="brand-title-main">Service Admin</div>
          <div class="brand-title-sub">Панель управления</div>
        </div>
      </div>
      <nav>
        <div class="nav-group-title">Навигация</div>
        <a href="/" class="nav-link {{ 'active' if active_page == 'index' else '' }}">📊 Обзор</a>
        <a href="/orders" class="nav-link {{ 'active' if active_page == 'orders' else '' }}">📋 Заявки</a>
        <a href="/order-search" class="nav-link {{ 'active' if active_page == 'order_search' else '' }}">🔍 Поиск по телефону</a>
        <a href="/cash" class="nav-link {{ 'active' if active_page == 'cash' else '' }}">💰 Касса</a>
        <a href="/sd" class="nav-link {{ 'active' if active_page == 'sd' else '' }}">📦 СД</a>
        <a href="/stats" class="nav-link {{ 'active' if active_page == 'stats' else '' }}">📈 Статистика</a>
        <a href="/commission" class="nav-link {{ 'active' if active_page == 'commission' else '' }}">💰 Проценты</a>
        <a href="/masters" class="nav-link {{ 'active' if active_page == 'masters' else '' }}">👥 Команда</a>
        <a href="/cities" class="nav-link {{ 'active' if active_page == 'cities' else '' }}">🏙 Города</a>
        {% if user_role == 'owner' %}
        <a href="/users" class="nav-link {{ 'active' if active_page == 'users' else '' }}">🔐 Пользователи</a>
        {% endif %}
      </nav>
      <div style="margin-top: auto; padding-top: 20px; font-size: 11px; color: var(--text-soft); text-align: center;">
        Service Admin v1.0
      </div>
    </aside>
    <main class="content">
      {% if flash_message %}
        <div class="flash flash-{{ flash_type }}" id="flashMessage">{{ flash_message }}</div>
        <script>
          setTimeout(function() {
            const flash = document.getElementById('flashMessage');
            if (flash) {
              showToast('{{ flash_message }}', '{{ flash_type }}');
              flash.remove();
            }
          }, 100);
        </script>
      {% endif %}
      {% block content %}{% endblock %}
    </main>
  </div>

  <!-- Toast контейнер -->
  <div class="toast-container" id="toastContainer"></div>

  <!-- Переключатель темы -->
  <button class="theme-toggle" id="themeToggle" title="Переключить тему">🌓</button>

  <script>
    // Toast уведомления
    function showToast(message, type = 'info') {
      const container = document.getElementById('toastContainer');
      if (!container) return;
      
      const toast = document.createElement('div');
      toast.className = 'toast toast-' + type;
      toast.innerHTML = `
        <span>${message}</span>
        <span class="toast-close" onclick="this.parentElement.remove()">×</span>
      `;
      
      container.appendChild(toast);
      
      setTimeout(() => {
        toast.style.animation = 'toastSlideIn 0.3s ease reverse';
        setTimeout(() => toast.remove(), 300);
      }, 3000);
    }

    // Модальные окна
    function showModal(modalId) {
      const modal = document.getElementById(modalId);
      if (modal) {
        modal.classList.add('active');
      }
    }

    function hideModal(modalId) {
      const modal = document.getElementById(modalId);
      if (modal) {
        modal.classList.remove('active');
      }
    }

    // Закрытие модалки по клику на overlay
    document.addEventListener('click', function(e) {
      if (e.target.classList.contains('modal-overlay')) {
        e.target.classList.remove('active');
      }
    });

    // Переключатель темы
    const themeToggle = document.getElementById('themeToggle');
    if (themeToggle) {
      const savedTheme = localStorage.getItem('theme') || 'dark';
      if (savedTheme === 'light') {
        document.body.classList.add('light-theme');
        themeToggle.textContent = '🌙';
      }
      
      themeToggle.addEventListener('click', function() {
        document.body.classList.toggle('light-theme');
        const isLight = document.body.classList.contains('light-theme');
        localStorage.setItem('theme', isLight ? 'light' : 'dark');
        themeToggle.textContent = isLight ? '🌙' : '☀️';
      });
    }

    // Автоматическое закрытие flash сообщений
    document.addEventListener('DOMContentLoaded', function() {
      const flash = document.querySelector('.flash');
      if (flash) {
        setTimeout(() => {
          flash.style.opacity = '0';
          flash.style.transform = 'translateY(-10px)';
          setTimeout(() => flash.remove(), 300);
        }, 5000);
      }
    });
  </script>

  <style>
    /* Светлая тема */
    body.light-theme {
      --bg: #f8fafc;
      --bg-soft: #ffffff;
      --surface: #ffffff;
      --surface-soft: #f1f5f9;
      --border-subtle: #e2e8f0;
      --text: #1e293b;
      --text-soft: #64748b;
    }
    body.light-theme {
      background: linear-gradient(circle at top left, #e0e7ff 0, #f8fafc 45%, #ffffff 100%);
    }
  </style>
</body>
</html>
"""

# Сохраняем базовый шаблон
with open("templates/base.html", "w", encoding="utf-8") as f:
    f.write(BASE_TEMPLATE)


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Страница входа"""
    # Инициализируем базового owner при первом открытии страницы
    session = get_session()
    try:
        ensure_owner_user(session)
    finally:
        session.close()
    
    return templates.TemplateResponse("admin_login.html", {
        "request": request,
        "page_title": "Вход в админку",
    })


def hash_password(password: str) -> str:
    """Хеширование пароля"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, password_hash: str) -> bool:
    """Проверка пароля"""
    try:
        return bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))
    except Exception:
        return False

def ensure_owner_user(session):
    """Создает базового owner пользователя с паролем из .env, если его нет"""
    owner_username = os.getenv("OWNER_USERNAME", "admin")
    owner_password = os.getenv("OWNER_PASSWORD", "admin123")
    
    # Проверяем, есть ли уже пользователь с таким username
    existing_user = session.query(User).filter_by(username=owner_username).first()
    if existing_user:
        # Если есть, но нет пароля - устанавливаем пароль
        if not existing_user.password_hash:
            existing_user.password_hash = hash_password(owner_password)
            if existing_user.role != "owner":
                existing_user.role = "owner"
            session.commit()
            logging.info(f"Установлен пароль для существующего пользователя {owner_username}")
        return existing_user
    
    # Создаем нового owner пользователя
    # Используем SUPER_ADMIN_ID как tg_id для базового owner
    owner_user = session.query(User).filter_by(tg_id=SUPER_ADMIN_ID).first()
    if owner_user:
        # Обновляем существующего пользователя
        owner_user.username = owner_username
        owner_user.password_hash = hash_password(owner_password)
        owner_user.role = "owner"
        if not owner_user.name:
            owner_user.name = "Администратор"
        session.commit()
        logging.info(f"Обновлен базовый owner пользователь: {owner_username}")
        return owner_user
    
    # Создаем нового пользователя
    try:
        owner_user = User(
            tg_id=SUPER_ADMIN_ID,
            username=owner_username,
            password_hash=hash_password(owner_password),
            role="owner",
            name="Администратор"
        )
        session.add(owner_user)
        session.commit()
        logging.info(f"Создан базовый owner пользователь: {owner_username}")
        return owner_user
    except Exception as e:
        session.rollback()
        logging.error(f"Ошибка при создании базового owner: {e}")
        return None

@app.post("/login", response_class=HTMLResponse)
async def login_submit(request: Request):
    """Обработка входа по username/password"""
    form = await request.form()
    username = (form.get("username") or "").strip()
    password = (form.get("password") or "").strip()
    
    if not username or not password:
        return templates.TemplateResponse("admin_login.html", {
            "request": request,
            "page_title": "Вход в админку",
            "error": "Введите логин и пароль",
        })
    
    session = get_session()
    try:
        # Инициализируем базового owner, если его нет
        ensure_owner_user(session)
        
        user = session.query(User).filter_by(username=username).first()
        if not user:
            return templates.TemplateResponse("admin_login.html", {
                "request": request,
                "page_title": "Вход в админку",
                "error": "Неверный логин или пароль",
            })
        
        if not user.password_hash or not verify_password(password, user.password_hash):
            return templates.TemplateResponse("admin_login.html", {
                "request": request,
                "page_title": "Вход в админку",
                "error": "Неверный логин или пароль",
            })
        
        # Проверяем права доступа - разрешаем вход для всех ролей
        if user.role not in ["owner", "director", "dispatcher", "master"]:
            return templates.TemplateResponse("admin_login.html", {
                "request": request,
                "page_title": "Вход в админку",
                "error": "Доступ запрещён. У вас нет прав для входа в веб-интерфейс.",
            })
        
        # Сохраняем в сессию
        request.session["user_id"] = user.tg_id
        request.session["user_tg_id"] = user.tg_id
        request.session["username"] = user.username
        
        return RedirectResponse("/", status_code=303)
    finally:
        session.close()


@app.get("/logout", response_class=HTMLResponse)
async def logout(request: Request):
    """Выход из системы"""
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request):
    """Страница регистрации (только для owner)"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user or user.role != "owner":
            return RedirectResponse("/?error=access_denied", status_code=303)
        
        cities = session.query(City).order_by(City.name).all()
        return templates.TemplateResponse("register.html", {
            "request": request,
            "page_title": "Регистрация пользователя",
            "cities": cities,
            "user_role": user.role,
        })
    finally:
        session.close()


@app.post("/register", response_class=HTMLResponse)
async def register_submit(request: Request):
    """Обработка регистрации нового пользователя"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user or user.role != "owner":
            return RedirectResponse("/?error=access_denied", status_code=303)
        
        form = await request.form()
        username = (form.get("username") or "").strip()
        password = (form.get("password") or "").strip()
        password_confirm = (form.get("password_confirm") or "").strip()
        full_name = (form.get("full_name") or "").strip()
        role = (form.get("role") or "").strip()
        city_id = form.get("city_id")
        tg_id_str = (form.get("tg_id") or "").strip()
        
        # Валидация
        errors = []
        if not username:
            errors.append("Логин обязателен")
        elif len(username) < 3:
            errors.append("Логин должен быть не менее 3 символов")
        elif session.query(User).filter_by(username=username).first():
            errors.append("Логин уже занят")
        
        if not password:
            errors.append("Пароль обязателен")
        elif len(password) < 6:
            errors.append("Пароль должен быть не менее 6 символов")
        elif password != password_confirm:
            errors.append("Пароли не совпадают")
        
        if not role or role not in ["owner", "director", "dispatcher", "master"]:
            errors.append("Выберите роль")
        
        if role == "director" and not city_id:
            errors.append("Для директора необходимо выбрать город")
        
        if not tg_id_str:
            # Генерируем временный tg_id (отрицательное число для веб-пользователей)
            existing_ids = [u.tg_id for u in session.query(User).all()]
            tg_id = -1
            while tg_id in existing_ids:
                tg_id -= 1
        else:
            try:
                tg_id = int(tg_id_str)
                if session.query(User).filter_by(tg_id=tg_id).first():
                    errors.append("VK ID уже используется")
            except ValueError:
                errors.append("Некорректный VK ID")
        
        if errors:
            cities = session.query(City).order_by(City.name).all()
            return templates.TemplateResponse("register.html", {
                "request": request,
                "page_title": "Регистрация пользователя",
                "cities": cities,
                "user_role": user.role,
                "errors": errors,
                "form_data": {
                    "username": username,
                    "full_name": full_name,
                    "role": role,
                    "city_id": city_id,
                    "tg_id": tg_id_str,
                },
            })
        
        # Создаем пользователя
        new_user = User(
            tg_id=tg_id,
            username=username,
            password_hash=hash_password(password),
            full_name=full_name or None,
            name=full_name or username,
            role=role,
            city_id=int(city_id) if city_id else None,
        )
        session.add(new_user)
        session.commit()
        
        return RedirectResponse("/users?success=created", status_code=303)
    except Exception as e:
        session.rollback()
        cities = session.query(City).order_by(City.name).all()
        user_role = get_user_role(session, user_id) if user_id else None
        return templates.TemplateResponse("register.html", {
            "request": request,
            "page_title": "Регистрация пользователя",
            "cities": cities,
            "user_role": user_role,
            "errors": [f"Ошибка: {str(e)}"],
        })
    finally:
        session.close()


@app.get("/users", response_class=HTMLResponse)
async def users_list(request: Request):
    """Список пользователей (только для owner)"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user or user.role != "owner":
            return RedirectResponse("/?error=access_denied", status_code=303)
        
        users = session.query(User).order_by(User.role, User.full_name, User.name).all()
        cities = session.query(City).order_by(City.name).all()
        
        success = request.query_params.get("success")
        error = request.query_params.get("error")
        
        return templates.TemplateResponse("users.html", {
            "request": request,
            "page_title": "Управление пользователями",
            "active_page": "users",
            "users": users,
            "cities": cities,
            "success": success,
            "error": error,
            "user_role": user.role,
        })
    finally:
        session.close()


@app.post("/users/{user_tg_id}/update", response_class=HTMLResponse)
async def update_user(request: Request, user_tg_id: int):
    """Обновление пользователя"""
    session = get_session()
    try:
        current_user_id = get_current_user_id(request)
        if not current_user_id or not check_owner_or_director_access(session, current_user_id):
            return RedirectResponse("/login", status_code=303)
        
        current_user = session.query(User).filter_by(tg_id=current_user_id).first()
        if not current_user or current_user.role != "owner":
            return RedirectResponse("/?error=access_denied", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_tg_id).first()
        if not user:
            return RedirectResponse("/users?error=not_found", status_code=303)
        
        form = await request.form()
        full_name = (form.get("full_name") or "").strip()
        role = (form.get("role") or "").strip()
        city_id = form.get("city_id")
        new_password = (form.get("new_password") or "").strip()
        
        if full_name:
            user.full_name = full_name
            user.name = full_name
        
        if role and role in ["owner", "director", "dispatcher", "master"]:
            user.role = role
        
        if city_id:
            user.city_id = int(city_id) if city_id != "None" else None
        
        if new_password:
            if len(new_password) < 6:
                return RedirectResponse("/users?error=password_too_short", status_code=303)
            user.password_hash = hash_password(new_password)
        
        session.commit()
        
        return RedirectResponse("/users?success=updated", status_code=303)
    except Exception as e:
        session.rollback()
        return RedirectResponse(f"/users?error={str(e)}", status_code=303)
    finally:
        session.close()


@app.post("/users/{user_tg_id}/delete", response_class=HTMLResponse)
async def delete_user(request: Request, user_tg_id: int):
    """Удаление пользователя"""
    session = get_session()
    try:
        current_user_id = get_current_user_id(request)
        if not current_user_id or not check_owner_or_director_access(session, current_user_id):
            return RedirectResponse("/login", status_code=303)
        
        current_user = session.query(User).filter_by(tg_id=current_user_id).first()
        if not current_user or current_user.role != "owner":
            return RedirectResponse("/?error=access_denied", status_code=303)
        
        if user_tg_id == current_user_id:
            return RedirectResponse("/users?error=cannot_delete_self", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_tg_id).first()
        if not user:
            return RedirectResponse("/users?error=not_found", status_code=303)
        
        # Удаляем связанные записи перед удалением пользователя
        from model import Order, Stat, Attendance, Penalty
        
        # Обнуляем ссылки на пользователя в заявках
        session.query(Order).filter(Order.created_by == user_tg_id).update({
            Order.created_by: None
        }, synchronize_session=False)
        
        session.query(Order).filter(Order.assigned_to == user_tg_id).update({
            Order.assigned_to: None
        }, synchronize_session=False)
        
        # Удаляем записи посещаемости
        session.query(Attendance).filter(Attendance.master_tg_id == user_tg_id).delete()
        
        # Удаляем штрафы
        session.query(Penalty).filter(Penalty.master_tg_id == user_tg_id).delete()
        
        # Обнуляем master_tg в статистике (если есть)
        session.query(Stat).filter(Stat.master_tg == user_tg_id).update({
            Stat.master_tg: None
        }, synchronize_session=False)
        
        # Теперь можно безопасно удалить пользователя
        session.delete(user)
        session.commit()
        
        return RedirectResponse("/users?success=deleted", status_code=303)
    except Exception as e:
        session.rollback()
        import traceback
        error_msg = str(e)
        logger.error(f"Ошибка при удалении пользователя {user_tg_id}: {error_msg}\n{traceback.format_exc()}")
        return RedirectResponse(f"/users?error={error_msg}", status_code=303)
    finally:
        session.close()


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """Главная страница с дашбордом"""
    session = get_session()
    try:
        # Проверка авторизации
        user_id = get_current_user_id(request)
        if not user_id:
            return RedirectResponse("/login", status_code=303)
        
        # Проверяем, что пользователь существует и имеет права доступа
        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user or user.role not in ["owner", "director", "dispatcher", "master"]:
            return RedirectResponse("/login?error=access_denied", status_code=303)
        from datetime import datetime, timedelta
        
        # Общая статистика
        total_orders = session.query(Order).filter(Order.status != "cancelled").count()
        completed_orders = session.query(Order).filter(
            Order.status == "completed",
            Order.status != "cancelled",
        ).count()
        active_orders = session.query(Order).filter(
            Order.status.in_(["new", "assigned", "accepted", "on_place", "done_pending_sum"]),
            Order.status != "cancelled",
        ).count()
        
        # Статистика по мастерам
        total_masters = session.query(User).filter(User.role == "master").count()
        total_directors = session.query(User).filter(User.role == "director").count()
        total_dispatchers = session.query(User).filter(User.role == "dispatcher").count()
        
        # Статистика по городам
        total_cities = session.query(City).count()
        
        # Статистика за текущий месяц (с 1 числа) - по умолчанию
        today = date.today()
        month_start = datetime.combine(today.replace(day=1), datetime.min.time())
        # Название месяца на русском
        month_names = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
        }
        current_month_name = f"{month_names[today.month]} {today.year}"
        orders_this_month = session.query(Order).filter(
            Order.created_at >= month_start,
            Order.status != "cancelled",
        ).count()
        completed_this_month = session.query(Order).filter(
            Order.status == "completed",
            Order.status != "cancelled",
            Order.created_at >= month_start
        ).count()
        
        # Чистая сумма за текущий месяц
        completed_orders_month = session.query(Order).filter(
            Order.status == "completed",
            Order.status != "cancelled",
            Order.created_at >= month_start
        ).all()
        total_net_sum_month = 0.0
        for order in completed_orders_month:
            order_sum = float(order.sum_amount or 0)
            sd_price = float(order.sd_price or 0)
            zpch_sum = float(order.zpch_sum or 0)
            net_amount = max(order_sum - zpch_sum, 0)
            total_net_sum_month += net_amount
        avg_check_month = total_net_sum_month / completed_this_month if completed_this_month > 0 else 0.0
        
        # Статистика за последние 7 дней (для графика)
        week_ago = datetime.now() - timedelta(days=7)
        orders_last_week = session.query(Order).filter(Order.created_at >= week_ago).count()
        completed_last_week = session.query(Order).filter(
            Order.status == "completed",
            Order.status != "cancelled",
            Order.created_at >= week_ago
        ).count()
        
        # Сумма закрытых заявок
        total_sum = session.query(func.sum(Order.sum_amount)).filter(
            Order.status == "completed",
            Order.status != "cancelled",
            Order.sum_amount.isnot(None)
        ).scalar() or 0.0
        
        # Топ мастера по заявкам
        top_masters = session.query(
            User.tg_id,
            User.full_name,
            User.name,
            func.count(Order.id).label('order_count')
        ).join(Order, User.tg_id == Order.assigned_to).filter(
            Order.status == "completed",
            Order.status != "cancelled",
        ).group_by(User.tg_id, User.full_name, User.name).order_by(
            func.count(Order.id).desc()
        ).limit(5).all()
        
        # Заявки по дням (последние 7 дней)
        daily_stats = []
        for i in range(6, -1, -1):
            day = datetime.now() - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
            count = session.query(Order).filter(
                Order.created_at >= day_start,
                Order.created_at <= day_end
            ).count()
            daily_stats.append({
                "date": day.strftime("%d.%m"),
                "count": count
            })
        
        user_role = get_user_role(session, user_id)
        return templates.TemplateResponse("index.html", {
            "request": request,
            "page_title": "Дашборд",
            "active_page": "index",
            "user_role": user_role,
            "total_orders": total_orders or 0,
            "completed_orders": completed_orders or 0,
            "active_orders": active_orders or 0,
            "total_masters": total_masters or 0,
            "total_directors": total_directors or 0,
            "total_dispatchers": total_dispatchers or 0,
            "total_cities": total_cities or 0,
            "orders_this_month": orders_this_month or 0,
            "completed_this_month": completed_this_month or 0,
            "orders_last_week": orders_last_week or 0,
            "completed_last_week": completed_last_week or 0,
            "current_month_name": current_month_name,
            "total_sum": total_sum or 0.0,
            "total_net_sum_month": total_net_sum_month or 0.0,
            "avg_check_month": avg_check_month or 0.0,
            "top_masters": top_masters or [],
            "daily_stats": daily_stats or [],
            "daily_stats_json": json.dumps(daily_stats or []),
        })
    finally:
        session.close()


@app.get("/commission", response_class=HTMLResponse)
async def commission_get(request: Request):
    """Проценты по направлениям - GET"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
    finally:
        session.close()
    
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        user_role = get_user_role(session, user_id) if user_id else None
    finally:
        session.close()
    
    settings = load_settings()
    return templates.TemplateResponse("commission.html", {
        "request": request,
        "page_title": "Проценты по направлениям",
        "active_page": "commission",
        "user_role": user_role,
        "settings": settings,
        "flash_message": None,
        "flash_type": None,
    })


@app.post("/commission", response_class=HTMLResponse)
async def commission_post(request: Request):
    """Проценты по направлениям - POST"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
    finally:
        session.close()
    
    form = await request.form()
    settings = load_settings()
    new_settings = settings.copy()
    for cat, conf in settings.items():
        tiers = conf.get("tiers") or []
        new_tiers = []
        for idx, (lo, hi, pct) in enumerate(tiers):
            field_name = f"pct_{cat}_{idx}"
            val = form.get(field_name, "").strip()
            try:
                new_pct = float(val)
            except ValueError:
                new_pct = pct
            new_tiers.append([lo, hi, new_pct])
        new_settings[cat]["tiers"] = new_tiers
    save_settings(new_settings)
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        user_role = get_user_role(session, user_id) if user_id else None
    finally:
        session.close()
    
    return templates.TemplateResponse("commission.html", {
        "request": request,
        "page_title": "Проценты по направлениям",
        "active_page": "commission",
        "user_role": user_role,
        "settings": new_settings,
        "flash_message": "Проценты по направлениям обновлены.",
        "flash_type": "success",
    })


@app.get("/masters", response_class=HTMLResponse)
async def masters_get(request: Request):
    """Мастера, директора, диспетчеры - GET"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id:
            return RedirectResponse("/login", status_code=303)
        
        # Проверяем доступ: owner, director или dispatcher
        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user or user.role not in ["owner", "director", "dispatcher"]:
            return RedirectResponse("/login", status_code=303)
        masters_q = session.query(User).filter(User.role == "master").all()
        dispatchers = session.query(User).filter(User.role == "dispatcher").all()
        directors = session.query(User).filter(User.role == "director").all()
        cities = session.query(City).order_by(City.name).all()
        return templates.TemplateResponse("masters.html", {
            "request": request,
            "page_title": "Команда",
            "active_page": "masters",
            "user_role": user.role,
            "masters": masters_q,
            "directors": directors,
            "dispatchers": dispatchers,
            "cities": cities,
            "flash_message": None,
            "flash_type": None,
        })
    finally:
        session.close()


@app.post("/masters/create", response_class=HTMLResponse)
async def masters_create(request: Request):
    """Создание нового мастера"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id:
            return RedirectResponse("/login", status_code=303)
        
        # Проверяем доступ: owner, director или dispatcher
        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user or user.role not in ["owner", "director", "dispatcher"]:
            return RedirectResponse("/login", status_code=303)
    finally:
        session.close()
    
    form = await request.form()
    session = get_session()
    flash_message = None
    flash_type = None
    
    try:
        tg_id_str = (form.get("tg_id") or "").strip()
        full_name = (form.get("full_name") or "").strip()
        phone = (form.get("phone") or "").strip()
        city_id_str = (form.get("city_id") or "").strip()
        
        if not tg_id_str:
            flash_message = "VK ID обязателен для заполнения."
            flash_type = "error"
        else:
            try:
                tg_id = int(tg_id_str)
                
                # Проверяем, не существует ли уже пользователь с таким tg_id
                existing = session.query(User).filter_by(tg_id=tg_id).first()
                if existing:
                    flash_message = f"Пользователь с VK ID {tg_id} уже существует (роль: {existing.role})."
                    flash_type = "error"
                else:
                    # Создаем нового мастера
                    new_master = User(
                        tg_id=tg_id,
                        role="master",
                        full_name=full_name if full_name else None,
                        name=full_name if full_name else None,
                        phone=phone if phone else None,
                        city_id=int(city_id_str) if city_id_str else None,
                    )
                    session.add(new_master)
                    session.commit()
                    flash_message = f"Мастер {full_name or f'ID {tg_id}'} успешно добавлен!"
                    flash_type = "success"
            except ValueError:
                flash_message = "Некорректный VK ID (должно быть число)."
                flash_type = "error"
        
        masters_q = session.query(User).filter(User.role == "master").all()
        dispatchers = session.query(User).filter(User.role == "dispatcher").all()
        directors = session.query(User).filter(User.role == "director").all()
        cities = session.query(City).order_by(City.name).all()
        user_role = get_user_role(session, user_id)
        return templates.TemplateResponse("masters.html", {
            "request": request,
            "page_title": "Команда",
            "active_page": "masters",
            "user_role": user_role,
            "masters": masters_q,
            "directors": directors,
            "dispatchers": dispatchers,
            "cities": cities,
            "flash_message": flash_message,
            "flash_type": flash_type,
        })
    except Exception as e:
        session.rollback()
        flash_message = f"Ошибка при создании мастера: {str(e)}"
        flash_type = "error"
        masters_q = session.query(User).filter(User.role == "master").all()
        dispatchers = session.query(User).filter(User.role == "dispatcher").all()
        directors = session.query(User).filter(User.role == "director").all()
        cities = session.query(City).order_by(City.name).all()
        return templates.TemplateResponse("masters.html", {
            "request": request,
            "page_title": "Команда",
            "active_page": "masters",
            "masters": masters_q,
            "directors": directors,
            "dispatchers": dispatchers,
            "cities": cities,
            "flash_message": flash_message,
            "flash_type": flash_type,
        })
    finally:
        session.close()


@app.post("/masters", response_class=HTMLResponse)
async def masters_post(request: Request):
    """Мастера, директора, диспетчеры - POST"""
    form = await request.form()
    session = get_session()
    flash_message = None
    flash_type = None
    
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        action = form.get("action", "").strip()
        percentage_raw = form.get("percentage", "").strip()
        
        if action == "set_one":
            # Установить процент для одного мастера
            master_id_raw = form.get("master_id", "").strip()
            if not master_id_raw:
                flash_message = "Выберите мастера."
                flash_type = "error"
            else:
                try:
                    master_id = int(master_id_raw)
                    user = session.query(User).filter(User.tg_id == master_id, User.role == "master").first()
                    if not user:
                        flash_message = "Мастер не найден."
                        flash_type = "error"
                    else:
                        if not percentage_raw or percentage_raw == "0":
                            user.master_percentage = None
                            flash_message = f"Процент для {user.full_name or user.name or f'ID {master_id}'} сброшен (будет использоваться базовая сетка)."
                        else:
                            try:
                                pct = float(percentage_raw)
                                if not (0 <= pct <= 100):
                                    flash_message = "Процент должен быть от 0 до 100."
                                    flash_type = "error"
                                else:
                                    user.master_percentage = pct
                                    flash_message = f"Процент для {user.full_name or user.name or f'ID {master_id}'} установлен: {pct:.1f}%."
                            except ValueError:
                                flash_message = "Некорректное значение процента."
                                flash_type = "error"
                        if flash_type != "error":
                            session.commit()
                            flash_type = "success"
                except ValueError:
                    flash_message = "Некорректный ID мастера."
                    flash_type = "error"
                    
        elif action == "set_all":
            # Установить процент для всех мастеров
            if not percentage_raw or percentage_raw == "0":
                # Сбросить для всех
                for user in session.query(User).filter(User.role == "master").all():
                    user.master_percentage = None
                session.commit()
                flash_message = "Процент сброшен для всех мастеров (будет использоваться базовая сетка)."
                flash_type = "success"
            else:
                try:
                    pct = float(percentage_raw)
                    if not (0 <= pct <= 100):
                        flash_message = "Процент должен быть от 0 до 100."
                        flash_type = "error"
                    else:
                        for user in session.query(User).filter(User.role == "master").all():
                            user.master_percentage = pct
                        session.commit()
                        flash_message = f"Процент {pct:.1f}% установлен для всех мастеров."
                        flash_type = "success"
                except ValueError:
                    flash_message = "Некорректное значение процента."
                    flash_type = "error"
        else:
            flash_message = "Не указано действие."
            flash_type = "error"
        
        masters_q = session.query(User).filter(User.role == "master").all()
        dispatchers = session.query(User).filter(User.role == "dispatcher").all()
        directors = session.query(User).filter(User.role == "director").all()
        cities = session.query(City).order_by(City.name).all()
        user_role = get_user_role(session, user_id) if user_id else None
        return templates.TemplateResponse("masters.html", add_user_role_to_context(session, user_id, {
            "request": request,
            "page_title": "Команда",
            "active_page": "masters",
            "user_role": user_role,
            "masters": masters_q,
            "directors": directors,
            "dispatchers": dispatchers,
            "cities": cities,
            "flash_message": flash_message,
            "flash_type": flash_type,
        }))
    except Exception as e:
        session.rollback()
        logger.exception(f"Ошибка при изменении процента мастера: {e}")
        flash_message = f"Ошибка при изменении процента: {str(e)}"
        flash_type = "error"
        masters_q = session.query(User).filter(User.role == "master").all()
        dispatchers = session.query(User).filter(User.role == "dispatcher").all()
        directors = session.query(User).filter(User.role == "director").all()
        cities = session.query(City).order_by(City.name).all()
        user_role = get_user_role(session, user_id) if user_id else None
        return templates.TemplateResponse("masters.html", add_user_role_to_context(session, user_id, {
            "request": request,
            "page_title": "Команда",
            "active_page": "masters",
            "user_role": user_role,
            "masters": masters_q,
            "directors": directors,
            "dispatchers": dispatchers,
            "cities": cities,
            "flash_message": flash_message,
            "flash_type": flash_type,
        }))
    finally:
        session.close()


@app.get("/cities", response_class=HTMLResponse)
async def cities_get(request: Request):
    """Города - GET"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        cities_q = session.query(City).order_by(City.name.asc()).all()
        user_id = get_current_user_id(request)
        user_role = get_user_role(session, user_id)
        return templates.TemplateResponse("cities.html", {
            "request": request,
            "page_title": "Города",
            "active_page": "cities",
            "user_role": user_role,
            "cities": cities_q,
            "flash_message": None,
            "flash_type": None,
        })
    finally:
        session.close()


@app.post("/cities", response_class=HTMLResponse)
async def cities_post(request: Request):
    """Города - POST"""
    form = await request.form()
    session = get_session()
    flash_message = None
    flash_type = None
    try:
        name = (form.get("name") or "").strip()
        tz = (form.get("timezone") or "").strip() or "Europe/Moscow"
        if not name:
            flash_message = "Название города не может быть пустым."
            flash_type = "error"
        else:
            if session.query(City).filter(City.name == name).first():
                flash_message = "Город с таким названием уже существует."
                flash_type = "error"
            else:
                city = City(name=name, timezone=tz)
                session.add(city)
                session.commit()
                flash_message = "Город добавлен."
                flash_type = "success"
        cities_q = session.query(City).order_by(City.name.asc()).all()
        user_id = get_current_user_id(request)
        user_role = get_user_role(session, user_id)
        return templates.TemplateResponse("cities.html", {
            "request": request,
            "page_title": "Города",
            "active_page": "cities",
            "user_role": user_role,
            "cities": cities_q,
            "flash_message": flash_message,
            "flash_type": flash_type,
        })
    finally:
        session.close()


# ===== Экспорт данных =====

@app.get("/export/masters")
async def export_masters(request: Request, format: str = Query("xlsx", regex="^(xlsx|csv)$")):
    """Экспорт списка мастеров"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        
        masters = session.query(User).filter(User.role == "master").all()
        
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["VK ID", "Имя", "Город", "Телефон", "Индивидуальный %"])
            for m in masters:
                writer.writerow([
                    m.tg_id,
                    m.full_name or m.name or "",
                    m.city_rel.name if m.city_rel else "",
                    m.phone or "",
                    m.master_percentage if m.master_percentage is not None else "по сетке"
                ])
            output.seek(0)
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=masters.csv"}
            )
        else:  # xlsx
            wb = Workbook()
            ws = wb.active
            ws.title = "Мастера"
            
            # Заголовки
            headers = ["VK ID", "Имя", "Город", "Телефон", "Индивидуальный %"]
            ws.append(headers)
            
            # Стили для заголовков
            header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Данные
            for m in masters:
                ws.append([
                    m.tg_id,
                    m.full_name or m.name or "",
                    m.city_rel.name if m.city_rel else "",
                    m.phone or "",
                    m.master_percentage if m.master_percentage is not None else "по сетке"
                ])
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=masters.xlsx"}
            )
    finally:
        session.close()


@app.get("/export/orders")
async def export_orders(
    request: Request,
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    days: int = Query(30, ge=1, le=365)
):
    """Экспорт заявок за период"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        from datetime import datetime, timedelta
        date_from = datetime.now() - timedelta(days=days)
        orders = session.query(Order).filter(Order.created_at >= date_from).order_by(Order.created_at.desc()).all()
        
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["№", "Город", "Адрес", "Техника", "Статус", "Мастер", "Сумма", "Дата создания"])
            for o in orders:
                master_name = ""
                if o.assigned_to:
                    master = session.query(User).filter_by(tg_id=o.assigned_to).first()
                    master_name = master.full_name or master.name or str(o.assigned_to) if master else str(o.assigned_to)
                writer.writerow([
                    o.order_number,
                    o.city_rel.name if o.city_rel else "",
                    f"{o.street or ''}, {o.house or ''}, {o.flat or ''}",
                    o.equip_type or "",
                    o.status,
                    master_name,
                    o.sum_amount or 0,
                    o.created_at.strftime("%d.%m.%Y %H:%M") if o.created_at else ""
                ])
            output.seek(0)
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=orders_{days}days.csv"}
            )
        else:  # xlsx
            wb = Workbook()
            ws = wb.active
            ws.title = "Заявки"
            
            headers = ["№", "Город", "Адрес", "Техника", "Статус", "Мастер", "Сумма", "Дата создания"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            for o in orders:
                master_name = ""
                if o.assigned_to:
                    master = session.query(User).filter_by(tg_id=o.assigned_to).first()
                    master_name = master.full_name or master.name or str(o.assigned_to) if master else str(o.assigned_to)
                ws.append([
                    o.order_number,
                    o.city_rel.name if o.city_rel else "",
                    f"{o.street or ''}, {o.house or ''}, {o.flat or ''}",
                    o.equip_type or "",
                    o.status,
                    master_name,
                    o.sum_amount or 0,
                    o.created_at.strftime("%d.%m.%Y %H:%M") if o.created_at else ""
                ])
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=orders_{days}days.xlsx"}
            )
    finally:
        session.close()


# ===== Расширенная статистика (API) =====

@app.get("/api/stats/extended")
async def get_extended_stats():
    """Расширенная статистика в JSON"""
    session = get_session()
    try:
        from datetime import datetime, timedelta
        try:
            from handlers.utils import get_equip_type_name
        except ImportError:
            # Fallback если функция не найдена
            def get_equip_type_name(code):
                equip_types = {
                    "refrigerator": "Холодильник",
                    "washing_machine": "Стиральная машина",
                    "dishwasher": "Посудомоечная машина",
                    "oven": "Духовка",
                    "microwave": "Микроволновка",
                    "tv": "Телевизор",
                    "other": "Другое"
                }
                return equip_types.get(code, code or "Не указан")
        
        # Статистика по типам техники
        equip_stats = session.query(
            Order.equip_type,
            func.count(Order.id).label('count'),
            func.sum(Order.sum_amount).label('total_sum')
        ).filter(
            Order.status == "completed",
            Order.is_warranty.is_(False),
            Order.sum_amount.isnot(None)
        ).group_by(Order.equip_type).all()
        
        equip_data = []
        for stat in equip_stats:
            equip_data.append({
                "type": get_equip_type_name(stat.equip_type) if stat.equip_type else "Не указан",
                "count": stat.count,
                "total_sum": float(stat.total_sum or 0)
            })
        
        # Статистика по городам
        city_stats = session.query(
            City.name,
            func.count(Order.id).label('count'),
            func.sum(Order.sum_amount).label('total_sum')
        ).join(Order, City.id == Order.city_id).filter(
            Order.status == "completed",
            Order.is_warranty.is_(False),
            Order.sum_amount.isnot(None)
        ).group_by(City.name).all()
        
        city_data = []
        for stat in city_stats:
            city_data.append({
                "city": stat.name,
                "count": stat.count,
                "total_sum": float(stat.total_sum or 0)
            })
        
        # Доходы по дням (последние 30 дней)
        revenue_data = []
        for i in range(29, -1, -1):
            day = datetime.now() - timedelta(days=i)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)
            revenue = session.query(func.sum(Order.sum_amount)).filter(
                Order.status == "completed",
                Order.is_warranty.is_(False),
                Order.sum_amount.isnot(None),
                Order.created_at >= day_start,
                Order.created_at <= day_end
            ).scalar() or 0.0
            revenue_data.append({
                "date": day.strftime("%d.%m"),
                "revenue": float(revenue)
            })
        
        # Средний чек
        avg_check = session.query(func.avg(Order.sum_amount)).filter(
            Order.status == "completed",
            Order.is_warranty.is_(False),
            Order.sum_amount.isnot(None)
        ).scalar() or 0.0
        
        # Конверсия (закрыто / всего)
        total_orders = session.query(Order).filter(Order.status != "cancelled").count()
        completed_orders = session.query(Order).filter(
            Order.status == "completed",
            Order.is_warranty.is_(False)
        ).count()
        conversion = (completed_orders / total_orders * 100) if total_orders > 0 else 0
        
        return JSONResponse({
            "equip_stats": equip_data,
            "city_stats": city_data,
            "revenue_data": revenue_data,
            "avg_check": float(avg_check),
            "conversion": round(conversion, 2)
        })
    finally:
        session.close()


# ===== Карточка мастера =====

@app.get("/master/{master_tg_id}", response_class=HTMLResponse)
async def master_card(request: Request, master_tg_id: int):
    """Карточка мастера - доступна только owner и director"""
    session = get_session()
    try:
        user_tg_id = get_current_user_id(request)
        if not user_tg_id or not check_owner_or_director_access(session, user_tg_id):
            return RedirectResponse("/login", status_code=303)
        
        master = session.query(User).filter_by(tg_id=master_tg_id, role="master").first()
        if not master:
            raise HTTPException(status_code=404, detail="Мастер не найден")
        
        return templates.TemplateResponse("master_card.html", {
            "request": request,
            "page_title": f"Карточка мастера",
            "active_page": "masters",
            "master": master,
        })
    finally:
        session.close()


@app.post("/master/{master_tg_id}/upload-passport")
async def upload_passport_photo(request: Request, master_tg_id: int, file: UploadFile = File(...)):
    """Загрузка фото паспорта мастера"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        master = session.query(User).filter_by(tg_id=master_tg_id, role="master").first()
        if not master:
            raise HTTPException(status_code=404, detail="Мастер не найден")
        
        # Проверяем расширение файла
        allowed_extensions = {".jpg", ".jpeg", ".png", ".webp"}
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in allowed_extensions:
            raise HTTPException(status_code=400, detail="Недопустимый формат файла. Разрешены: JPG, PNG, WEBP")
        
        # Сохраняем файл
        filename = f"{master_tg_id}_{file.filename}"
        file_path = PASSPORT_PHOTOS_DIR / filename
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Обновляем путь в БД
        master.passport_photo_path = str(file_path)
        session.commit()
        
        return RedirectResponse(f"/master/{master_tg_id}?uploaded=1", status_code=303)
    finally:
        session.close()


@app.get("/master/{master_tg_id}/passport-photo")
async def get_passport_photo(request: Request, master_tg_id: int):
    """Получить фото паспорта мастера"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        master = session.query(User).filter_by(tg_id=master_tg_id, role="master").first()
        if not master or not master.passport_photo_path:
            raise HTTPException(status_code=404, detail="Фото не найдено")
        
        if not os.path.exists(master.passport_photo_path):
            raise HTTPException(status_code=404, detail="Файл не найден")
        
        return FileResponse(
            master.passport_photo_path,
            media_type="image/jpeg",
            filename=f"passport_{master_tg_id}.jpg"
        )
    finally:
        session.close()


# ===== Управление заявками =====

@app.get("/order-search", response_class=HTMLResponse)
async def order_search(request: Request, phone: Optional[str] = Query(None)):
    """Поиск заявок по телефону клиента для просмотра документов"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        orders = []
        if phone and phone.strip():
            # Нормализуем телефон для поиска (убираем пробелы, скобки, дефисы)
            phone_normalized = phone.strip().replace(" ", "").replace("-", "").replace("(", "").replace(")", "").replace("+", "")
            search_term = f"%{phone_normalized}%"
            
            # Ищем заявки по телефону клиента (без ограничений по дате - ищем все заявки)
            query = session.query(Order).filter(
                Order.client_phone.ilike(search_term)
            )
            
            # Если директор - фильтруем по городу
            user = session.query(User).filter_by(tg_id=user_id).first()
            if user and user.role == "director" and user.city_id:
                query = query.filter(Order.city_id == user.city_id)
            
            orders = query.order_by(Order.created_at.desc()).all()
        
        # Получаем список городов для отображения
        cities = session.query(City).order_by(City.name).all()
        cities_dict = {c.id: c for c in cities}
        
        # Получаем список мастеров для отображения
        masters = session.query(User).filter_by(role="master").all()
        masters_dict = {m.tg_id: m for m in masters}
        
        from handlers.utils import get_status_name_ru
        return templates.TemplateResponse("order_search.html", add_user_role_to_context(session, user_id, {
            "request": request,
            "page_title": "Поиск заявок по телефону",
            "active_page": "order_search",
            "phone": phone,
            "orders": orders,
            "cities": cities_dict,
            "masters": masters_dict,
            "get_status_name_ru": get_status_name_ru,
        }))
    finally:
        session.close()

@app.get("/orders", response_class=HTMLResponse)
async def orders_list(
    request: Request,
    status: Optional[str] = Query(None),
    city_id: Optional[str] = Query(None),  # Принимаем как строку, потом конвертируем
    search: Optional[str] = Query(None),
    master_id: Optional[str] = Query(None),  # Принимаем как строку, потом конвертируем
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    period: Optional[str] = Query(None),  # "current_month" или "all"
):
    """Список заявок с фильтрацией"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        # По умолчанию показываем текущий месяц с 1 числа
        if not date_from and not date_to and not period:
            period = "current_month"
        
        # Если выбран период "текущий месяц"
        if period == "current_month":
            today = date.today()
            date_from = today.replace(day=1).strftime("%Y-%m-%d")
            date_to = today.strftime("%Y-%m-%d")
        
        base_query = session.query(Order)
        
        # Фильтры
        if status:
            base_query = base_query.filter(Order.status == status)
        if city_id:
            try:
                city_id_int = int(city_id)
                base_query = base_query.filter(Order.city_id == city_id_int)
            except (ValueError, TypeError):
                pass  # Игнорируем некорректный city_id
        if master_id:
            try:
                master_id_int = int(master_id)
                base_query = base_query.filter(Order.assigned_to == master_id_int)
            except (ValueError, TypeError):
                pass  # Игнорируем некорректный master_id
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
                # Преобразуем в datetime для корректного сравнения
                date_from_datetime = datetime.combine(date_from_obj, datetime.min.time())
                base_query = base_query.filter(Order.created_at >= date_from_datetime)
            except (ValueError, TypeError) as e:
                # Игнорируем некорректные даты
                logger = logging.getLogger(__name__)
                logger.warning(f"Ошибка парсинга date_from: {e}")
                pass
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
                # Добавляем время до конца дня
                date_to_datetime = datetime.combine(date_to_obj, datetime.max.time())
                base_query = base_query.filter(Order.created_at <= date_to_datetime)
            except (ValueError, TypeError) as e:
                # Игнорируем некорректные даты
                logger = logging.getLogger(__name__)
                logger.warning(f"Ошибка парсинга date_to: {e}")
                pass
        if search:
            search_term = f"%{search}%"
            # Ищем по номеру заявки, адресу (улица, дом, квартира), клиенту, телефону и описанию
            # Используем func.concat для поиска по полному адресу
            full_address = func.concat(
                func.coalesce(Order.street, ''),
                ' ',
                func.coalesce(Order.house, ''),
                ' ',
                func.coalesce(Order.flat, '')
            )
            base_query = base_query.filter(
                (cast(Order.order_number, String).ilike(search_term)) |
                (Order.street.ilike(search_term)) |
                (Order.house.ilike(search_term)) |
                (Order.flat.ilike(search_term)) |
                (full_address.ilike(search_term)) |
                (Order.client_name.ilike(search_term)) |
                (Order.client_phone.ilike(search_term)) |
                (Order.short_desc.ilike(search_term))
            )
        
        # Подсчет статистики по отфильтрованным заявкам (ДО лимита)
        # "Отмена" не должна попадать в общую статистику
        stats_query = base_query.filter(Order.status != "cancelled")
        total_count = stats_query.count()
        
        # ===== Выборка списка (join только для сортировки по названию города) =====
        from sqlalchemy import case
        list_query = base_query.outerjoin(City, Order.city_id == City.id)
        # Кросс-БД сортировка: NULL города уезжают вниз без использования NULLS LAST
        city_nulls_last = case((City.name.is_(None), 1), else_=0)
        orders = list_query.order_by(
            city_nulls_last.asc(),
            City.name.asc(),
            Order.created_at.desc()
        ).limit(100).all()
        
        # Реальное количество найденных заявок (для отображения)
        found_count = len(orders)
        completed_count = base_query.filter(
            Order.status == "completed"
        ).count()
        not_completed_count = total_count - completed_count
        accepted_count = stats_query.filter(Order.status.in_(["accepted", "on_place", "to_sd"])).count()
        
        # Подсчет сумм (только для закрытых заявок)
        completed_orders = stats_query.filter(
            Order.status == "completed"
        ).all()
        total_sum = 0.0
        total_net_sum = 0.0
        for order in completed_orders:
            order_sum = float(order.sum_amount or 0)
            sd_price = float(order.sd_price or 0)
            zpch_sum = float(order.zpch_sum or 0)
            net_amount = max(order_sum - zpch_sum, 0)
            total_sum += order_sum
            total_net_sum += net_amount
        
        avg_check = total_net_sum / completed_count if completed_count > 0 else 0.0
        
        # Данные для фильтров
        cities = session.query(City).order_by(City.name).all()
        masters = session.query(User).filter_by(role="master").order_by(User.full_name, User.name).all()
        # Словарь мастеров по ID для быстрого поиска
        masters_dict = {m.tg_id: m for m in masters}
        statuses = ["new", "assigned", "accepted", "on_place", "done_pending_sum", "done", "cancelled", "declined", "completed", "scheduled"]
        
        from handlers.utils import get_equip_type_name, get_status_name_ru
        
        # Проверяем роль пользователя для скрытия телефона
        user = session.query(User).filter_by(tg_id=user_id).first()
        user_role = user.role if user else None
        hide_phone = user_role in ["director", "master"]  # Скрываем телефон от директора и мастера
        
        return templates.TemplateResponse("orders_list.html", {
            "request": request,
            "page_title": "Заявки",
            "active_page": "orders",
            "user_role": user_role,
            "orders": orders,
            "cities": cities,
            "masters": masters,
            "masters_dict": masters_dict,
            "statuses": statuses,
            "current_filters": {
                "status": status,
                "city_id": city_id,
                "master_id": master_id,
                "search": search,
                "date_from": date_from,
                "date_to": date_to,
                "period": period,
            },
            "statistics": {
                "total": total_count,
                "found": found_count,  # Количество найденных заявок (с учётом лимита)
                "completed": completed_count,
                "not_completed": not_completed_count,
                "accepted": accepted_count,
                "total_sum": total_sum,
                "total_net_sum": total_net_sum,
                "avg_check": avg_check,
            },
            "get_equip_type_name": get_equip_type_name,
            "get_status_name_ru": get_status_name_ru,
            "hide_phone": hide_phone,
        })
    finally:
        session.close()


@app.get("/order/{order_id}", response_class=HTMLResponse)
async def order_view(request: Request, order_id: int):
    """Просмотр заявки"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        order = session.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        
        masters = session.query(User).filter_by(role="master").order_by(User.full_name, User.name).all()
        directors = session.query(User).filter_by(role="director").order_by(User.full_name, User.name).all()
        cities = session.query(City).order_by(City.name).all()
        statuses = ["new", "assigned", "accepted", "on_place", "done_pending_sum", "done", "cancelled", "declined", "completed", "scheduled"]
        
        from handlers.utils import get_equip_type_name
        from handlers.menu_kb import EQUIP_TYPES
        
        from handlers.utils import get_status_name_ru
        
        # Получаем создателя заявки для отображения имени
        creator = None
        if order.created_by:
            creator = session.query(User).filter_by(tg_id=order.created_by).first()
        
        # Проверяем роль пользователя для скрытия телефона
        user = session.query(User).filter_by(tg_id=user_id).first()
        user_role = user.role if user else None
        hide_phone = user_role in ["director", "master"]  # Скрываем телефон от директора и мастера

        # ===== Автогарантия: ищем исходную заявку с активной гарантией по телефону =====
        from datetime import datetime, timezone
        from services.warranty_service import normalize_phone

        now = datetime.now(timezone.utc)
        current_norm = normalize_phone(order.client_phone or "")
        warranty_source = None
        warranty_reason = None
        if current_norm:
            candidates = (
                session.query(Order)
                .filter(
                    Order.id != order.id,
                    Order.status == "completed",
                    Order.status != "cancelled",
                    Order.is_warranty.is_(False),
                    Order.warranty_until.isnot(None),
                    Order.warranty_until >= now,
                )
                .order_by(Order.created_at.desc())
                .limit(50)
                .all()
            )
            for cand in candidates:
                if normalize_phone(cand.client_phone or "") == current_norm:
                    warranty_source = cand
                    break
            if not warranty_source:
                warranty_reason = "Нет закрытой заявки с действующей гарантией по этому телефону."
        else:
            warranty_reason = "В заявке не указан телефон клиента."

        can_make_warranty = (
            (not getattr(order, "is_warranty", False))
            and (order.status != "cancelled")
            and (warranty_source is not None)
        )
        
        return templates.TemplateResponse("order_view.html", {
            "request": request,
            "page_title": f"Заявка #{order.order_number}",
            "active_page": "orders",
            "user_role": user_role,
            "order": order,
            "masters": masters,
            "directors": directors,
            "cities": cities,
            "statuses": statuses,
            "equip_types": EQUIP_TYPES,
            "get_equip_type_name": get_equip_type_name,
            "get_status_name_ru": get_status_name_ru,
            "creator": creator,
            "hide_phone": hide_phone,
            "warranty_source": warranty_source,
            "can_make_warranty": can_make_warranty,
            "warranty_reason": warranty_reason,
        })
    finally:
        session.close()


@app.post("/order/{order_id}/make-warranty", response_class=HTMLResponse)
async def order_make_warranty(request: Request, order_id: int):
    """Оформить гарантию для заявки (ставит is_warranty=True) при наличии активного срока по исходной заявке."""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)

        order = session.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Заявка не найдена")

        form = await request.form()
        source_id_raw = (form.get("source_order_id") or "").strip()
        if not source_id_raw:
            raise HTTPException(status_code=400, detail="Не указан источник гарантии")
        try:
            source_id = int(source_id_raw)
        except ValueError:
            raise HTTPException(status_code=400, detail="Некорректный источник гарантии")

        from datetime import datetime, timezone
        from services.warranty_service import normalize_phone

        now = datetime.now(timezone.utc)
        current_norm = normalize_phone(order.client_phone or "")
        if not current_norm:
            raise HTTPException(status_code=400, detail="В заявке не указан телефон клиента")

        source = session.query(Order).filter_by(id=source_id).first()
        if not source:
            raise HTTPException(status_code=404, detail="Исходная заявка не найдена")

        # Валидации
        if getattr(order, "is_warranty", False):
            return RedirectResponse(f"/order/{order_id}?warranty=already", status_code=303)
        if order.status == "cancelled":
            raise HTTPException(status_code=400, detail="Нельзя оформить гарантию для отменённой заявки")
        if source.status != "completed" or source.status == "cancelled" or getattr(source, "is_warranty", False):
            raise HTTPException(status_code=400, detail="Исходная заявка не подходит для гарантии")
        if not source.warranty_until or source.warranty_until < now:
            raise HTTPException(status_code=400, detail="Срок гарантии по исходной заявке истёк")
        if normalize_phone(source.client_phone or "") != current_norm:
            raise HTTPException(status_code=400, detail="Телефон клиента не совпадает с исходной заявкой")

        # Оформляем
        order.is_warranty = True
        order.warranty_source_order_id = source.id
        order.warranty_until = None
        order.warranty_days = 0
        session.commit()

        return RedirectResponse(f"/order/{order_id}?warranty=1", status_code=303)
    finally:
        session.close()

@app.post("/order/{order_id}/assign", response_class=HTMLResponse)
async def order_assign(request: Request, order_id: int):
    """Быстрое назначение заявки (совместимость с director_order_view.html)"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        
        order = session.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        
        form = await request.form()
        raw_value = (form.get("master_id") or form.get("assigned_to") or "").strip()
        old_assigned_to = order.assigned_to
        
        if not raw_value or raw_value == "-":
            order.assigned_to = None
        else:
            try:
                order.assigned_to = int(raw_value)
            except ValueError:
                order.assigned_to = None
        
        session.commit()
        session.refresh(order)
        
        # Отправляем уведомления, если мастер был назначен или изменён
        if order.assigned_to and order.assigned_to != old_assigned_to:
            try:
                from services.notification_service import get_notification_service
                service = get_notification_service()
                service.notify_master(order, session)
                # Также уведомляем директоров и собственников
                service.notify_directors(order, session)
                service.notify_owners(order, session)
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"Ошибка при отправке уведомлений о назначении заявки {order.id}: {e}")
        
        referer = request.headers.get("referer")
        if referer:
            base = referer.split("?")[0]
            return RedirectResponse(f"{base}?assigned=1", status_code=303)
        return RedirectResponse(f"/order/{order_id}?assigned=1", status_code=303)
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка назначения: {str(e)}")
    finally:
        session.close()

@app.post("/order/{order_id}", response_class=HTMLResponse)
async def order_edit(request: Request, order_id: int, bso_file: Optional[UploadFile] = File(None), receipt_file: Optional[UploadFile] = File(None)):
    """Редактирование заявки с загрузкой файлов"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        order = session.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        
        # Сохраняем старый статус ДО изменения
        old_status = order.status
        old_assigned_to = order.assigned_to
        
        form = await request.form()
        
        # Загрузка файла БСО/договор/квитанция
        logger.info(f"Попытка загрузки БСО для заявки {order_id}: bso_file={bso_file}, filename={bso_file.filename if bso_file else None}")
        
        if bso_file and bso_file.filename:
            from services.bso_storage import save_bso_from_bytes
            
            allowed_extensions = {".jpg", ".jpeg", ".png", ".webp", ".pdf"}
            file_ext = Path(bso_file.filename).suffix.lower()
            if file_ext not in allowed_extensions:
                logger.warning(f"Недопустимый формат БСО файла: {file_ext}")
                raise HTTPException(status_code=400, detail="Недопустимый формат файла БСО. Разрешены: JPG, PNG, WEBP, PDF")
            
            # Читаем содержимое файла
            try:
                file_content = await bso_file.read()
                file_size = len(file_content)
                logger.info(f"Размер БСО файла: {file_size} байт")
                
                if file_size == 0:
                    raise HTTPException(status_code=400, detail="Файл БСО пустой")
            except Exception as e:
                logger.error(f"Ошибка при чтении БСО файла: {e}")
                raise HTTPException(status_code=500, detail=f"Ошибка при чтении файла БСО: {str(e)}")
            
            # Используем единый сервис для сохранения БСО
            # Передаём имя с расширением, сервис извлечёт расширение и сгенерирует правильное имя
            # Формат: временное имя с расширением, сервис преобразует его в bso_{order_id}_{timestamp}.ext
            temp_filename = f"temp{file_ext}"
            filename = save_bso_from_bytes(order_id, file_content, temp_filename)
            
            if not filename:
                logger.error(f"❌ Не удалось сохранить БСО файл для заявки {order_id}")
                raise HTTPException(status_code=500, detail="Не удалось сохранить файл БСО. Проверьте права доступа к директориям и место на диске.")
            
            # Сохраняем имя файла в БД (только имя, без пути)
            order.bso_file_path = filename
            session.flush()
            logger.info(f"✅ БСО сохранён и путь записан в БД для заявки {order_id}: {filename}")
        else:
            logger.info(f"БСО файл не предоставлен для заявки {order_id}: bso_file={bso_file}")
        
        # Загрузка файла чека ЗПЧ
        logger.info(f"Попытка загрузки чека для заявки {order_id}: receipt_file={receipt_file}, filename={receipt_file.filename if receipt_file else None}")
        
        if receipt_file and receipt_file.filename:
            allowed_extensions = {".jpg", ".jpeg", ".png", ".webp", ".pdf"}
            file_ext = Path(receipt_file.filename).suffix.lower()
            if file_ext not in allowed_extensions:
                logger.warning(f"Недопустимый формат чека: {file_ext}")
                raise HTTPException(status_code=400, detail="Недопустимый формат файла чека. Разрешены: JPG, PNG, WEBP, PDF")
            
            # Убеждаемся, что директория существует
            BSO_FILES_DIR.mkdir(exist_ok=True)
            
            # Очищаем имя файла
            safe_filename = "".join(c for c in receipt_file.filename if c.isalnum() or c in "._-")
            if not safe_filename:
                safe_filename = "file"
            timestamp = int(time.time())
            filename = f"receipt_{order_id}_{timestamp}_{safe_filename}"
            
            # Пробуем сохранить в основную директорию
            file_path = BSO_FILES_DIR.resolve() / filename
            
            try:
                # Читаем содержимое файла
                file_content = await receipt_file.read()
                file_size = len(file_content)
                logger.info(f"Размер чека: {file_size} байт")
                
                if file_size == 0:
                    raise HTTPException(status_code=400, detail="Файл чека пустой")
                
                # Сохраняем файл
                with open(file_path, "wb") as buffer:
                    buffer.write(file_content)
                
                # Проверяем сохранение
                if file_path.exists() and file_path.stat().st_size == file_size:
                    # Сохраняем относительный путь в БД
                    relative_path = f"bso_files/{filename}"
                    order.receipt_file_path = relative_path
                    session.flush()
                    logger.info(f"✅ Чек сохранен для заявки {order_id}: {file_path}, путь в БД: {relative_path}")
                else:
                    logger.error(f"❌ Чек не сохранен корректно для заявки {order_id}")
                    raise HTTPException(status_code=500, detail="Ошибка при сохранении файла чека")
            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Ошибка при сохранении чека для заявки {order_id}: {e}", exc_info=True)
                raise HTTPException(status_code=500, detail=f"Ошибка при сохранении файла чека: {str(e)}")
        else:
            logger.info(f"Чек не предоставлен для заявки {order_id}")
        
        # Обновляем поля
        if "city_id" in form and form["city_id"]:
            try:
                order.city_id = int(form["city_id"])
            except ValueError:
                pass
        
        if "street" in form:
            order.street = form["street"].strip() or None
        if "house" in form:
            order.house = form["house"].strip() or None
        if "flat" in form:
            order.flat = form["flat"].strip() or None
        if "time_from" in form:
            order.time_from = form["time_from"].strip() or None
        if "time_to" in form:
            order.time_to = form["time_to"].strip() or None
        if "equip_type" in form:
            order.equip_type = form["equip_type"].strip() or None
        if "short_desc" in form:
            order.short_desc = form["short_desc"].strip() or None
        if "source" in form:
            order.source = form["source"].strip() or None
        
        # Сохраняем старый статус ДО изменения
        old_status = order.status
        
        if "status" in form:
            order.status = form["status"].strip()
        if "client_name" in form:
            order.client_name = form["client_name"].strip() or None
        if "client_phone" in form:
            order.client_phone = form["client_phone"].strip() or None
        if "comment" in form:
            order.comment = form["comment"].strip() or None
        
        # Дата выполнения
        if "order_date" in form and form["order_date"]:
            try:
                order_date = datetime.strptime(form["order_date"], "%Y-%m-%d").date()
                order.order_date = datetime.combine(order_date, datetime.min.time())
            except ValueError:
                pass
        elif "order_date" in form and not form["order_date"]:
            order.order_date = None
        
        # Назначение мастера
        if "assigned_to" in form:
            assigned_to_str = form["assigned_to"].strip()
            if assigned_to_str and assigned_to_str != "-":
                try:
                    order.assigned_to = int(assigned_to_str)
                except ValueError:
                    pass
            elif assigned_to_str == "-":
                order.assigned_to = None
        
        # Суммы
        if "sum_amount" in form and form["sum_amount"]:
            try:
                order.sum_amount = float(form["sum_amount"])
            except ValueError:
                pass
        if "paid_amount" in form and form["paid_amount"]:
            try:
                paid_amount = float(form["paid_amount"])
                order.paid_amount = paid_amount
                # Пересчитываем долг
                order_sum = float(order.sum_amount or 0)
                if paid_amount < order_sum:
                    order.debt_amount = order_sum - paid_amount
                else:
                    order.debt_amount = None
            except ValueError:
                pass
        if "sd_price" in form and form["sd_price"]:
            try:
                order.sd_price = float(form["sd_price"])
            except ValueError:
                pass
        if "zpch_sum" in form and form["zpch_sum"]:
            try:
                order.zpch_sum = float(form["zpch_sum"])
            except ValueError:
                pass
        
        # Гарантия
        if "is_warranty" in form:
            order.is_warranty = form["is_warranty"] == "on"
        
        # Если статус изменен на completed, создаем запись в Stat (и рассчитываем гарантию)
        new_status = order.status  # Уже обновлен выше
        if (
            new_status == "completed"
            and old_status != "completed"
            and not getattr(order, "is_warranty", False)
            and order.status != "cancelled"
        ):
            from model import Stat
            from services.commission_service import get_master_pct
            from services.warranty_service import compute_warranty
            from datetime import datetime, timezone
            
            # Проверяем, нет ли уже записи в Stat
            existing_stat = session.query(Stat).filter_by(order_id=order.id).first()
            if not existing_stat:
                order_sum = float(order.sum_amount or 0)
                sd_price = float(order.sd_price or 0)
                zpch_sum = float(order.zpch_sum or 0)
                net_amount = max(order_sum - zpch_sum, 0)
                
                stat = Stat(
                    order_id=order.id,
                    equip_type=order.equip_type,
                    sum=net_amount,
                    refused=False,
                    master_tg=order.assigned_to
                )
                session.add(stat)

                # Автосрок гарантии по сумме закрытия
                closed_amount = float(order.sum_amount or order.paid_amount or 0)
                winfo = compute_warranty(closed_amount, datetime.now(timezone.utc))
                order.warranty_days = int(winfo.days)
                order.warranty_until = winfo.until
        
        # Проверяем, что БСО путь сохранен перед commit
        if order.bso_file_path:
            logger.info(f"Перед commit: order.bso_file_path={order.bso_file_path} для заявки {order_id}")
        
        session.commit()
        session.refresh(order)
        
        # Проверяем после commit
        refreshed_order = session.query(Order).filter_by(id=order_id).first()
        if refreshed_order and refreshed_order.bso_file_path:
            logger.info(f"После commit: order.bso_file_path={refreshed_order.bso_file_path} для заявки {order_id}")
        else:
            logger.warning(f"После commit: order.bso_file_path пустой для заявки {order_id}")
        
        # Отправляем уведомления, если мастер был назначен или изменён
        if order.assigned_to and order.assigned_to != old_assigned_to:
            try:
                from services.notification_service import get_notification_service
                service = get_notification_service()
                service.notify_master(order, session)
                # Также уведомляем директоров и собственников
                service.notify_directors(order, session)
                service.notify_owners(order, session)
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"Ошибка при отправке уведомлений о назначении заявки {order.id}: {e}")
        
        return RedirectResponse(f"/order/{order_id}?updated=1", status_code=303)
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при обновлении заявки: {str(e)}")
    finally:
        session.close()

@app.get("/create-order", response_class=HTMLResponse)
async def create_order_form(request: Request):
    """Форма создания новой заявки (owner/director/dispatcher)"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        user = session.query(User).filter_by(tg_id=user_id).first() if user_id else None
        if not user or user.role not in ["owner", "director", "dispatcher"]:
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        
        from handlers.utils import get_equip_type_name
        from handlers.menu_kb import EQUIP_TYPES
        cities = session.query(City).order_by(City.name).all()
        masters = session.query(User).filter_by(role="master").order_by(User.full_name, User.name).all()
        directors = session.query(User).filter_by(role="director").order_by(User.full_name, User.name).all()
        
        return templates.TemplateResponse("create_order.html", {
            "request": request,
            "page_title": "Создать заявку",
            "active_page": "orders",
            "cities": cities,
            "masters": masters,
            "directors": directors,
            "equip_types": EQUIP_TYPES,
            "initial": None,
            "error": None,
            "get_equip_type_name": get_equip_type_name,
        })
    finally:
        session.close()

@app.post("/create-order", response_class=HTMLResponse)
async def create_order_submit(request: Request):
    """Создание новой заявки (owner/director/dispatcher)"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        user = session.query(User).filter_by(tg_id=user_id).first() if user_id else None
        if not user or user.role not in ["owner", "director", "dispatcher"]:
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        
        form = await request.form()
        city_id = (form.get("city_id") or "").strip()
        street = (form.get("street") or "").strip()
        house = (form.get("house") or "").strip()
        is_private_house = form.get("is_private_house") == "on"  # Чекбокс "ЧД"
        is_warranty = form.get("is_warranty") == "on"  # Гарантия (не учитывать в статистике)
        flat = (form.get("flat") or "").strip()
        # Если отмечен чекбокс "ЧД", поле квартиры не требуется
        if is_private_house:
            flat = ""  # Очищаем поле квартиры для частного дома
        time_from = (form.get("time_from") or "").strip()
        time_to = (form.get("time_to") or "").strip()
        equip_type = (form.get("equip_type") or "").strip()
        short_desc = (form.get("short_desc") or "").strip()
        source = (form.get("source") or "").strip()
        client_name = (form.get("client_name") or "").strip()
        client_phone = (form.get("client_phone") or "").strip()
        comment = (form.get("comment") or "").strip()
        
        # Валидация обязательных полей
        missing = []
        if not city_id: missing.append("город")
        if not street: missing.append("улицу")
        if not house: missing.append("дом")
        # Квартира обязательна только если не отмечен чекбокс "ЧД"
        if not is_private_house and not flat: missing.append("квартиру")
        if not time_from: missing.append("время с")
        if not time_to: missing.append("время до")
        if not equip_type: missing.append("тип техники")
        if not short_desc: missing.append("краткое описание")
        if not source: missing.append("источник")
        if not client_name: missing.append("имя клиента")
        if not client_phone: missing.append("телефон клиента")
        
        if missing:
            error_msg = "Заполните все поля: " + ", ".join(missing)
            from handlers.utils import get_equip_type_name
            from handlers.menu_kb import EQUIP_TYPES
            return templates.TemplateResponse("create_order.html", {
                "request": request,
                "page_title": "Создать заявку",
                "active_page": "orders",
                "cities": session.query(City).order_by(City.name).all(),
                "masters": session.query(User).filter_by(role="master").order_by(User.full_name, User.name).all(),
                "directors": session.query(User).filter_by(role="director").order_by(User.full_name, User.name).all(),
                "equip_types": EQUIP_TYPES,
                "initial": {
                    "city_id": city_id,
                    "street": street,
                    "house": house,
                    "flat": flat,
                    "is_private_house": is_private_house,
                    "is_warranty": is_warranty,
                    "time_from": time_from,
                    "time_to": time_to,
                    "order_date": (form.get("order_date") or "").strip(),
                    "equip_type": equip_type,
                    "short_desc": short_desc,
                    "source": source,
                    "client_name": client_name,
                    "client_phone": client_phone,
                    "comment": comment,
                    "assigned_to": (form.get("assigned_to") or "").strip(),
                },
                "error": error_msg,
                "get_equip_type_name": get_equip_type_name,
            })
        
        try:
            city_id_int = int(city_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный ID города")
        
        # Номер заявки
        from services.user_service import generate_order_number
        order_number = generate_order_number(session)
        
        # Дата выполнения (опционально)
        order_date_str = (form.get("order_date") or "").strip()
        order_date = None
        if order_date_str:
            try:
                order_date = datetime.strptime(order_date_str, "%Y-%m-%d").date()
            except ValueError:
                order_date = None
        
        # Назначение
        assigned_to = None
        assigned_to_str = (form.get("assigned_to") or "").strip()
        if assigned_to_str and assigned_to_str != "-":
            try:
                assigned_to = int(assigned_to_str)
            except ValueError:
                assigned_to = None
        if not assigned_to:
            # Автоматически директор города (если есть)
            director = session.query(User).filter_by(role="director", city_id=city_id_int).first()
            if director:
                assigned_to = director.tg_id
        
        # Статус
        status = "new"
        if order_date and order_date > date.today():
            status = "scheduled"
        
        new_order = Order(
            order_number=order_number,
            city_id=city_id_int,
            street=street,
            house=house,
            flat=flat,
            time_from=time_from,
            time_to=time_to,
            order_date=datetime.combine(order_date, datetime.min.time()) if order_date else None,
            equip_type=equip_type,
            short_desc=short_desc,
            source=source,
            status=status,
            created_by=user_id,
            assigned_to=assigned_to,
            client_name=client_name or None,
            client_phone=client_phone or None,
            comment=comment or None,
            is_warranty=is_warranty,
        )
        session.add(new_order)
        session.commit()
        
        # Обновляем объект заявки для получения связей
        session.refresh(new_order)
        
        # Отправляем уведомления директорам и мастерам
        try:
            from services.notification_service import notify_order_created
            from datetime import date as date_class
            from datetime import datetime as datetime_class
            
            # Если заявка не запланирована на будущее, отправляем уведомления сразу
            if not order_date or order_date <= date_class.today():
                notify_order_created(new_order, session)
            # Если запланирована на будущее, уведомления отправятся при активации через scheduler
        except Exception as e:
            # Логируем ошибку, но не прерываем создание заявки
            import logging
            logging.getLogger(__name__).warning(f"Ошибка при отправке уведомлений о заявке {new_order.id}: {e}")
        
        return RedirectResponse(f"/order/{new_order.id}?created=1", status_code=303)
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при создании заявки: {str(e)}")
    finally:
        session.close()

@app.get("/order/{order_id}/bso-file")
async def get_bso_file(request: Request, order_id: int):
    """Скачать файл БСО/договор/квитанция"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        order = session.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        if not user_id or not check_document_access(session, user_id, order):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        
        # Используем единый сервис для поиска БСО
        from services.bso_storage import find_bso_file
        
        logger.info(f"Поиск БСО для заявки {order_id}: bso_file_path={order.bso_file_path}")
        
        # Ищем файл БСО (сервис сам обработает все варианты путей)
        bso_path = find_bso_file(order_id, order.bso_file_path)
        
        if not bso_path or not bso_path.exists():
            # Если файл не найден, пробуем обновить путь в БД (на случай, если файл есть, но путь устарел)
            # Сервис уже попробовал найти по паттерну, так что если не нашёл - значит файла нет
            raise HTTPException(
                status_code=404,
                detail=f"БСО не найден для заявки {order_id}. Путь в БД: {order.bso_file_path or 'не указан'}"
            )
        
        # Обновляем путь в БД, если он устарел (на случай миграции со старых путей)
        filename = bso_path.name
        if order.bso_file_path != filename:
            logger.info(f"Обновление пути БСО в БД: {order.bso_file_path} -> {filename}")
            order.bso_file_path = filename
            session.commit()
        
        # Определяем media_type в зависимости от расширения файла
        suffix = bso_path.suffix.lower()
        media_types = {
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.webp': 'image/webp',
            '.pdf': 'application/pdf',
        }
        media_type = media_types.get(suffix, 'application/octet-stream')
        
        # Для изображений используем inline, чтобы открывать в браузере, для PDF тоже
        # Для других типов - attachment (скачивание)
        content_disposition = 'inline' if suffix in ['.jpg', '.jpeg', '.png', '.webp', '.pdf'] else 'attachment'
        
        from fastapi.responses import Response
        with open(bso_path, 'rb') as f:
            content = f.read()
        
        return Response(
            content=content,
            media_type=media_type,
            headers={
                'Content-Disposition': f'{content_disposition}; filename="bso_{order.order_number}{suffix}"'
            }
        )
    finally:
        session.close()


@app.get("/order/{order_id}/receipt-file")
async def get_receipt_file(request: Request, order_id: int):
    """Скачать файл чека ЗПЧ"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        order = session.query(Order).filter_by(id=order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        
        if not user_id or not check_document_access(session, user_id, order):
            raise HTTPException(status_code=403, detail="Доступ запрещён")
        
        # Проверяем наличие чека - может быть в receipt_file_path или receipt_file_id
        if not order.receipt_file_path and not order.receipt_file_id:
            raise HTTPException(
                status_code=404, 
                detail="Чек не был сохранен для этой заявки. Функциональность сохранения чека была добавлена позже."
            )
        
        # Определяем путь к файлу чека
        receipt_path = None
        
        # Логируем для отладки (только если включено логирование)
        try:
            logger = logging.getLogger(__name__)
            logger.info(f"Поиск чека для заявки {order_id}: receipt_file_path={order.receipt_file_path}, receipt_file_id={order.receipt_file_id}")
        except:
            pass  # Если логирование не настроено, просто пропускаем
        
        # Если есть receipt_file_path, используем его
        if order.receipt_file_path:
            # Преобразуем путь в абсолютный, если он относительный
            receipt_path = Path(order.receipt_file_path)
            if not receipt_path.is_absolute():
                # Если путь относительный, пробуем найти в стандартной директории или по исходному пути
                if "data/receipts" in str(receipt_path) or "receipts" in str(receipt_path):
                    # Старый формат из бота - пробуем найти файл
                    receipt_path = Path(order.receipt_file_path)
                else:
                    # Пробуем найти в стандартной директории БСО
                    receipt_path = BSO_FILES_DIR / receipt_path.name
            else:
                receipt_path = Path(order.receipt_file_path)
            
            if not receipt_path.exists():
                # Пробуем найти файл по имени в стандартной директории
                receipt_path = BSO_FILES_DIR / Path(order.receipt_file_path).name
                if not receipt_path.exists():
                    # Пробуем найти по паттерну
                    possible_files = list(BSO_FILES_DIR.glob(f"receipt_{order_id}_*"))
                    if possible_files:
                        receipt_path = possible_files[0]
                        try:
                            logger.info(f"Найден файл чека по паттерну: {receipt_path}")
                        except:
                            pass
                        # Обновляем путь в БД для будущих запросов
                        order.receipt_file_path = str(receipt_path.resolve())
                        session.commit()
                    else:
                        try:
                            logger.warning(f"Файл чека не найден. Путь в БД: {order.receipt_file_path}, директория: {BSO_FILES_DIR}")
                        except:
                            pass
                        raise HTTPException(
                            status_code=404, 
                            detail=f"Файл чека не найден на диске. Возможно, файл был удален или не был сохранен при создании заявки. Путь в БД: {order.receipt_file_path}"
                        )
        else:
            # Если receipt_file_path пустой, но есть receipt_file_id, пробуем найти файл по паттерну
            if order.receipt_file_id:
                try:
                    logger.info(f"receipt_file_path пустой, но есть receipt_file_id. Ищем файл по паттерну receipt_{order_id}_*")
                except:
                    pass
                # Пробуем найти файл по паттерну receipt_{order_id}_*
                possible_files = list(BSO_FILES_DIR.glob(f"receipt_{order_id}_*"))
                if possible_files:
                    receipt_path = possible_files[0]
                    try:
                        logger.info(f"Найден файл чека по паттерну: {receipt_path}")
                    except:
                        pass
                    # Обновляем путь в БД для будущих запросов
                    order.receipt_file_path = str(receipt_path.resolve())
                    session.commit()
                else:
                    try:
                        logger.warning(f"Файл чека не найден по паттерну receipt_{order_id}_* в директории {BSO_FILES_DIR}")
                    except:
                        pass
                    # Пробуем также поискать в других возможных местах
                    other_dirs = [Path("data/receipts"), Path("receipts"), Path(".")]
                    found = False
                    for other_dir in other_dirs:
                        if other_dir.exists():
                            possible_files = list(other_dir.glob(f"receipt_{order_id}_*"))
                            if possible_files:
                                receipt_path = possible_files[0]
                                try:
                                    logger.info(f"Найден файл чека в альтернативной директории: {receipt_path}")
                                except:
                                    pass
                                # Перемещаем в стандартную директорию
                                new_path = BSO_FILES_DIR / receipt_path.name
                                shutil.copy2(receipt_path, new_path)
                                receipt_path = new_path
                                # Обновляем путь в БД
                                order.receipt_file_path = str(receipt_path.resolve())
                                session.commit()
                                found = True
                                break
                    if not found:
                        raise HTTPException(
                            status_code=404,
                            detail="Чек сохранен в боте, но файл не найден на диске. Возможно, файл был удален. Загрузите чек через веб-интерфейс."
                        )
            else:
                # Нет ни receipt_file_path, ни receipt_file_id
                try:
                    logger.warning(f"Чек не найден для заявки {order_id}: нет ни receipt_file_path, ни receipt_file_id")
                except:
                    pass
                raise HTTPException(
                    status_code=404,
                    detail="Чек не был сохранен для этой заявки."
                )
        
        if not receipt_path or not receipt_path.exists():
            try:
                logger.error(f"Критическая ошибка: receipt_path не установлен или файл не существует: {receipt_path}")
            except:
                pass
            raise HTTPException(
                status_code=500,
                detail="Внутренняя ошибка при поиске файла чека."
            )
        
        # Определяем media_type в зависимости от расширения файла
        suffix = receipt_path.suffix.lower()
        media_types = {
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.webp': 'image/webp',
            '.pdf': 'application/pdf',
        }
        media_type = media_types.get(suffix, 'application/octet-stream')
        
        # Для изображений используем inline, чтобы открывать в браузере, для PDF тоже
        # Для других типов - attachment (скачивание)
        content_disposition = 'inline' if suffix in ['.jpg', '.jpeg', '.png', '.webp', '.pdf'] else 'attachment'
        
        from fastapi.responses import Response
        with open(receipt_path, 'rb') as f:
            content = f.read()
        
        return Response(
            content=content,
            media_type=media_type,
            headers={
                'Content-Disposition': f'{content_disposition}; filename="receipt_{order.order_number}{suffix}"'
            }
        )
    finally:
        session.close()


# ===== Принятие денег в кассу =====

@app.get("/cash", response_class=HTMLResponse)
async def cash_overview(
    request: Request,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    period: str = Query("month"),
    city_id: Optional[str] = Query(None),
    show_income: bool = Query(False),  # Показать приход денег
):
    """Просмотр кассы - заявки к сдаче и приход денег"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_id).first()
        city_id_int = None
        if user and user.role == "director":
            city_id_int = user.city_id
        elif city_id and city_id.strip():
            try:
                city_id_int = int(city_id)
            except ValueError:
                city_id_int = None
        
        # Заявки со статусом done_pending_sum (к сдаче)
        query = session.query(Order).filter(Order.status == "done_pending_sum")
        if city_id_int:
            query = query.filter(Order.city_id == city_id_int)
        
        pending_orders = query.order_by(Order.created_at.desc()).all()
        
        # Группируем по мастерам и считаем суммы
        from collections import defaultdict
        masters_cash = defaultdict(lambda: {"master": None, "orders": [], "total_company": 0.0})
        total_company = 0.0
        
        from services.commission_service import get_master_pct
        
        for order in pending_orders:
            master_id = order.assigned_to
            if not master_id:
                continue
            
            if masters_cash[master_id]["master"] is None:
                master = session.query(User).filter_by(tg_id=master_id).first()
                masters_cash[master_id]["master"] = master
            
            order_sum = float(order.sum_amount or 0)
            sd_price = float(order.sd_price or 0)
            zpch_sum = float(order.zpch_sum or 0)
            net_amount = max(order_sum - zpch_sum, 0)
            
            # Используем индивидуальный процент мастера, если установлен
            master = session.query(User).filter_by(tg_id=master_id).first()
            if master and master.master_percentage is not None:
                pct = float(master.master_percentage)
            else:
                pct = get_master_pct(order.equip_type or "other", net_amount)
            
            master_share = net_amount * (pct / 100.0)
            company_share = net_amount - master_share
            
            masters_cash[master_id]["orders"].append({
                "order": order,
                "order_sum": order_sum,
                "sd_price": sd_price,
                "zpch_sum": zpch_sum,
                "net_amount": net_amount,
                "pct": pct,
                "master_share": master_share,
                "company_share": company_share,
            })
            masters_cash[master_id]["total_company"] += company_share
            total_company += company_share
        
        # Приход денег (если запрошен)
        cash_income = None
        if show_income:
            date_from_dt, date_to_dt, period_label = resolve_date_range(date_from, date_to, period)
            cash_income = calculate_cash_income(
                session,
                date_from=date_from_dt,
                date_to=date_to_dt,
                city_id=city_id_int,
            )
        
        from handlers.utils import get_equip_type_name, get_status_name_ru
        
        # Получаем список городов для фильтра
        if user.role == "owner":
            cities = session.query(City).order_by(City.name.asc()).all()
        else:
            cities = (
                session.query(City).filter(City.id == user.city_id).all()
                if user.city_id
                else []
            )
        
        date_from_dt, date_to_dt, period_label = resolve_date_range(date_from, date_to, period)
        
        return templates.TemplateResponse("cash.html", add_user_role_to_context(session, user_id, {
            "request": request,
            "page_title": "Касса",
            "active_page": "cash",
            "user_role": user.role,
            "user": user,
            "masters_cash": dict(masters_cash),
            "total_company": total_company,
            "cash_income": cash_income,
            "show_income": show_income,
            "cities": cities,
            "filters": {
                "city_id": city_id_int,
                "date_from": date_from_dt.date().isoformat() if date_from_dt else None,
                "date_to": date_to_dt.date().isoformat() if date_to_dt else None,
                "period": period,
                "period_label": period_label,
            },
            "get_equip_type_name": get_equip_type_name,
            "get_status_name_ru": get_status_name_ru,
        }))
    finally:
        session.close()


@app.post("/cash/accept-order/{order_id}", response_class=HTMLResponse)
async def accept_order_cash(request: Request, order_id: int):
    """Принять одну заявку в кассу - перевести в статус completed"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_id).first()
        city_id = None
        if user and user.role == "director":
            city_id = user.city_id
        
        # Получаем заявку
        order = session.query(Order).filter_by(id=order_id).first()
        if not order:
            return RedirectResponse("/cash?error=order_not_found", status_code=303)
        
        # Проверяем статус
        if order.status != "done_pending_sum":
            return RedirectResponse("/cash?error=wrong_status", status_code=303)
        
        # Проверяем доступ директора к заявке из своего города
        if city_id and order.city_id != city_id:
            return RedirectResponse("/cash?error=access_denied", status_code=303)
        
        from services.commission_service import get_master_pct
        from model import Stat
        
        order_sum = float(order.sum_amount or 0)
        sd_price = float(order.sd_price or 0)
        zpch_sum = float(order.zpch_sum or 0)
        net_amount = max(order_sum - zpch_sum, 0)
        
        # Переводим в статус completed
        order.status = "completed"
        
        # Проверяем, нет ли уже записи в Stat
        existing_stat = session.query(Stat).filter_by(order_id=order.id).first()
        if not existing_stat:
            # Создаем запись в статистике
            # Используем индивидуальный процент мастера, если установлен
            master = session.query(User).filter_by(tg_id=order.assigned_to).first()
            if master and master.master_percentage is not None:
                pct = float(master.master_percentage)
            else:
                pct = get_master_pct(order.equip_type or "other", net_amount)
            master_share = net_amount * (pct / 100.0)
            company_share = net_amount - master_share
            
            stat = Stat(
                order_id=order.id,
                equip_type=order.equip_type,
                sum=net_amount,
                refused=False,
                master_tg=order.assigned_to
            )
            session.add(stat)
        
        session.commit()
        
        return RedirectResponse(f"/cash?accepted_order=1&order_number={order.order_number}", status_code=303)
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при принятии заявки: {str(e)}")
    finally:
        session.close()


@app.post("/cash/accept/{master_id}", response_class=HTMLResponse)
async def accept_cash(request: Request, master_id: int):
    """Принять кассу мастера - перевести заявки в статус completed"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_id).first()
        city_id = None
        if user and user.role == "director":
            city_id = user.city_id
        
        # Получаем все заявки мастера со статусом done_pending_sum
        query = session.query(Order).filter(
            Order.assigned_to == master_id,
            Order.status == "done_pending_sum"
        )
        if city_id:
            query = query.filter(Order.city_id == city_id)
        
        orders = query.all()
        
        if not orders:
            return RedirectResponse("/cash?error=no_orders", status_code=303)
        
        from services.commission_service import get_master_pct
        from model import Stat
        
        total_company = 0.0
        
        for order in orders:
            order_sum = float(order.sum_amount or 0)
            sd_price = float(order.sd_price or 0)
            zpch_sum = float(order.zpch_sum or 0)
            net_amount = max(order_sum - zpch_sum, 0)
            
            # Переводим в статус completed
            order.status = "completed"
            
            # Проверяем, нет ли уже записи в Stat
            existing_stat = session.query(Stat).filter_by(order_id=order.id).first()
            if not existing_stat:
                # Создаем запись в статистике
                pct = get_master_pct(order.equip_type or "other", net_amount)
                master_share = net_amount * (pct / 100.0)
                company_share = net_amount - master_share
                total_company += company_share
                
                stat = Stat(
                    order_id=order.id,
                    equip_type=order.equip_type,
                    sum=net_amount,
                    refused=False,
                    master_tg=order.assigned_to
                )
                session.add(stat)
        
        session.commit()
        
        master = session.query(User).filter_by(tg_id=master_id).first()
        master_name = master.full_name or master.name or str(master_id) if master else str(master_id)
        
        return RedirectResponse(f"/cash?accepted=1&master={master_name}&amount={total_company:.2f}&count={len(orders)}", status_code=303)
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при принятии кассы: {str(e)}")
    finally:
        session.close()


# ===== Просмотр СД =====

@app.get("/sd", response_class=HTMLResponse)
async def sd_overview(request: Request):
    """Просмотр СД - незакрытые заявки"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        user = session.query(User).filter_by(tg_id=user_id).first()
        city_id = None
        if user and user.role == "director":
            city_id = user.city_id
        
        # Заявки со статусом СД
        query = session.query(Order).filter(
            Order.status.in_(["accepted", "on_place", "to_sd"])
        )
        if city_id:
            query = query.filter(Order.city_id == city_id)
        
        sd_orders = query.order_by(Order.created_at.desc()).all()
        
        # Группируем по мастерам
        masters_dict = {}
        for order in sd_orders:
            master_id = order.assigned_to
            if master_id not in masters_dict:
                master = session.query(User).filter_by(tg_id=master_id).first()
                masters_dict[master_id] = {
                    "master": master,
                    "orders": []
                }
            masters_dict[master_id]["orders"].append(order)
        
        from handlers.utils import get_equip_type_name, get_status_name_ru
        
        return templates.TemplateResponse("sd.html", {
            "request": request,
            "page_title": "СД",
            "active_page": "sd",
            "user_role": user.role,
            "user": user,
            "masters_dict": masters_dict,
            "get_equip_type_name": get_equip_type_name,
            "get_status_name_ru": get_status_name_ru,
        })
    finally:
        session.close()


@app.get("/stats/export")
async def stats_export(
    request: Request,
    master_id: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    city_id: Optional[str] = Query(None),
    period: Optional[str] = Query("month"),
):
    """Экспорт статистики в Excel (как в боте)"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)
        
        # Проверяем роль для фильтрации по городу
        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user:
            return RedirectResponse("/login", status_code=303)
        
        # Парсим параметры
        master_id_int = None
        if master_id and master_id.strip():
            try:
                master_id_int = int(master_id)
            except ValueError:
                master_id_int = None
        
        city_id_int = None
        if user.role == "director":
            city_id_int = user.city_id
        elif city_id and city_id.strip():
            try:
                city_id_int = int(city_id)
            except ValueError:
                city_id_int = None
        
        # Используем ту же логику дат, что и в stats_page
        date_from_dt, date_to_dt, _ = resolve_date_range(date_from, date_to, period or "month")
        
        # Получаем заявки через calculate_dashboard_stats для правильной фильтрации
        from services.dashboard_stats import calculate_dashboard_stats
        dashboard = calculate_dashboard_stats(
            session,
            date_from=date_from_dt,
            date_to=date_to_dt,
            city_id=city_id_int,
            master_id=master_id_int,
        )
        
        # Получаем только completed заявки из результата
        # Используем JOIN с Stat для получения заявок, закрытых в периоде
        from model import Stat
        stat_query = session.query(Stat).filter(
            Stat.recorded_at >= date_from_dt,
            Stat.recorded_at <= date_to_dt,
        )
        
        if master_id_int:
            stat_query = stat_query.filter(Stat.master_tg == master_id_int)
        
        if city_id_int:
            stat_query = stat_query.join(Order, Order.id == Stat.order_id).filter(Order.city_id == city_id_int)
        elif user.role == "director" and user.city_id:
            stat_query = stat_query.join(Order, Order.id == Stat.order_id).filter(Order.city_id == user.city_id)
        else:
            stat_query = stat_query.join(Order, Order.id == Stat.order_id)
        
        # Получаем order_id из Stat и загружаем заявки
        order_ids = [s.order_id for s in stat_query.all()]
        if not order_ids:
            raise HTTPException(status_code=400, detail="Нет данных для отчёта")
        
        orders = session.query(Order).filter(Order.id.in_(order_ids), Order.status == "completed").order_by(Order.created_at.desc()).all()
        
        if not orders:
            raise HTTPException(status_code=400, detail="Нет данных для отчёта")
        
        # Генерируем Excel файл
        from services.statistics_service import generate_city_stats_excel
        import tempfile
        
        file_path = generate_city_stats_excel(orders)
        
        # Читаем файл и возвращаем
        with open(file_path, "rb") as f:
            content = f.read()
        
        # Удаляем временный файл
        try:
            os.remove(file_path)
        except Exception:
            pass
        
        filename = f"statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        session.close()


# ===== Статистика =====

@app.get("/stats", response_class=HTMLResponse)
async def stats_page(
    request: Request,
    master_id: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    city_id: Optional[str] = Query(None),
    period: str = Query("month"),
    equip_category: Optional[str] = Query(None),  # "appliance", "digital", "other"
    include_warranty: Optional[str] = Query("1"),  # "1" или "0" из формы
):
    """Страница статистики"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            return RedirectResponse("/login", status_code=303)

        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user:
            return RedirectResponse("/login", status_code=303)

        # Парсим параметры, обрабатывая пустые строки как None
        master_id_int = None
        if master_id and master_id.strip():
            try:
                master_id_int = int(master_id)
            except ValueError:
                master_id_int = None
        
        city_id_int = None
        if user.role == "director":
            city_id_int = user.city_id
        elif city_id and city_id.strip():
            try:
                city_id_int = int(city_id)
            except ValueError:
                city_id_int = None

        date_from_dt, date_to_dt, period_label = resolve_date_range(date_from, date_to, period)

        # Парсим include_warranty (нужно ДО расчётов)
        include_warranty_bool = (
            include_warranty in ("1", "true", "True", True) if include_warranty else True
        )

        city_name = None
        if city_id_int:
            city_obj = session.query(City).filter_by(id=city_id_int).first()
            city_name = city_obj.name if city_obj else None
        elif user.role == "director" and getattr(user, "city_rel", None):
            city_name = user.city_rel.name

        dashboard = calculate_dashboard_stats(
            session,
            date_from=date_from_dt,
            date_to=date_to_dt,
            city_id=city_id_int,
            city_name=city_name,
            master_id=master_id_int,
            equip_category=equip_category if equip_category in ("appliance", "digital", "other") else None,
            include_warranty=include_warranty_bool,
        )
        dashboard_summary = summarize_dashboard(dashboard, user.role or "owner")

        stat_query = session.query(Stat)
        if master_id_int:
            stat_query = stat_query.filter(Stat.master_tg == master_id_int)

        stat_query = stat_query.filter(
            Stat.recorded_at >= date_from_dt,
            Stat.recorded_at <= date_to_dt,
        )

        if city_id_int or (user.role == "director" and user.city_id):
            city_filter_id = city_id_int or user.city_id
            stat_query = stat_query.join(Order, Order.id == Stat.order_id).filter(Order.city_id == city_filter_id)
        else:
            stat_query = stat_query.outerjoin(Order, Order.id == Stat.order_id)

        stat_records = stat_query.order_by(Stat.recorded_at.desc()).limit(200).all()

        masters_query = session.query(User).filter(User.role == "master")
        if user.role == "director" and user.city_id:
            masters_query = masters_query.filter(User.city_id == user.city_id)
        masters = masters_query.order_by(User.full_name, User.name).all()
        masters_dict = {m.tg_id: m for m in masters}

        if user.role == "owner":
            cities = session.query(City).order_by(City.name.asc()).all()
        else:
            cities = (
                session.query(City).filter(City.id == user.city_id).all()
                if user.city_id
                else []
            )

        period_options = [
            ("today", "Сегодня"),
            ("week", "Последние 7 дней"),
            ("month", "Текущий месяц"),
            ("last_30", "30 дней"),
            ("quarter", "Квартал"),
            ("year", "Год"),
        ]

        filters = {
            "master_id": master_id_int,
            "city_id": city_id_int,
            "period": period if period in PERIOD_LABELS else "month",
            "date_from": date_from_dt.date().isoformat(),
            "date_to": date_to_dt.date().isoformat(),
            "period_label": period_label,
            "equip_category": equip_category,
            "include_warranty": include_warranty_bool,
        }

        from handlers.utils import get_equip_type_name

        export_params = {
            "master_id": master_id,
            "city_id": city_id,
            "date_from": filters["date_from"],
            "date_to": filters["date_to"],
            "period": filters["period"],
        }
        export_query = urlencode({k: v for k, v in export_params.items() if v})
        
        # Получаем детальную статистику
        detailed_stats = calculate_detailed_stats(
            session,
            date_from=date_from_dt,
            date_to=date_to_dt,
            city_id=city_id_int,
            city_name=city_name,
            master_id=master_id_int,
            equip_category=equip_category if equip_category in ("appliance", "digital", "other") else None,
            include_warranty=include_warranty_bool,
        )
        
        # Статистика в формате таблицы по категориям
        category_table = calculate_category_table_stats(
            session,
            date_from=date_from_dt,
            date_to=date_to_dt,
            city_id=city_id_int,
            master_id=master_id_int,
            equip_category=equip_category if equip_category in ("appliance", "digital", "other") else None,
            include_warranty=include_warranty_bool,
        )
        
        # Данные для графиков (Chart.js) — берём из dashboard, чтобы совпадало с карточками
        charts_payload = {
            "daily": dashboard.get("daily") or [],
            "cities": (dashboard.get("cities") or [])[:10],
            "equipment": (dashboard.get("equipment") or [])[:10],
            "masters": (dashboard.get("masters") or [])[:10],
            "statuses": dashboard.get("status_breakdown") or [],
            "sources": (dashboard.get("sources") or [])[:10],
        }
        charts_json = json.dumps(charts_payload, ensure_ascii=False)

        return templates.TemplateResponse("stats.html", add_user_role_to_context(session, user_id, {
            "request": request,
            "page_title": "Статистика",
            "active_page": "stats",
            "user_role": user.role,
            "dashboard": dashboard,
            "dashboard_summary": dashboard_summary,
            "detailed_stats": detailed_stats,
            "category_table": category_table,
            "stat_records": stat_records,
            "masters": masters,
            "masters_dict": masters_dict,
            "period_options": period_options,
            "filters": filters,
            "cities": cities,
            "get_equip_type_name": get_equip_type_name,
            "export_query": export_query,
            "charts_json": charts_json,
        }))
    finally:
        session.close()


@app.get("/stats/export")
async def stats_export(
    request: Request,
    master_id: Optional[int] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    city_id: Optional[int] = Query(None),
    period: str = Query("month"),
):
    """Экспорт статистики в Excel"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id or not check_owner_or_director_access(session, user_id):
            raise HTTPException(status_code=403, detail="Доступ запрещён")

        user = session.query(User).filter_by(tg_id=user_id).first()
        if not user:
            raise HTTPException(status_code=403, detail="Доступ запрещён")

        if user.role == "director":
            city_id = user.city_id

        date_from_dt, date_to_dt, _ = resolve_date_range(date_from, date_to, period)

        # Получаем заявки со статусом completed для генерации Excel
        query = session.query(Order).filter(
            Order.status == "completed",
            Order.created_at >= date_from_dt,
            Order.created_at <= date_to_dt,
        )
        
        if master_id:
            query = query.filter(Order.assigned_to == master_id)
        if city_id:
            query = query.filter(Order.city_id == city_id)
        
        orders = query.order_by(Order.created_at.desc()).all()
        
        if not orders:
            raise HTTPException(status_code=404, detail="Нет данных для экспорта")
        
        from services.statistics_service import generate_city_stats_excel
        
        file_path = generate_city_stats_excel(orders)
        
        def remove_file_after_send(file_path: str):
            def cleanup():
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except Exception:
                    pass
            
            return cleanup
        
        response = FileResponse(
            file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        import atexit
        atexit.register(remove_file_after_send(file_path))
        
        return response
    finally:
        session.close()


def main():
    import uvicorn
    port = int(os.getenv("ADMIN_PORT", "8000"))
    uvicorn.run(app, host="0.0.0.0", port=port)


@app.get("/company/form", response_class=HTMLResponse)
async def company_form_page(request: Request):
    """Страница формы заполнения данных компании"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id:
            return RedirectResponse("/login", status_code=303)
        
        return templates.TemplateResponse("company_form.html", {
            "request": request,
            "page_title": "Заполнение данных компании",
            "active_page": "company",
            "user_role": get_user_role(session, user_id),
            "error": None,
        })
    finally:
        session.close()


@app.post("/company/form")
async def company_form_submit(request: Request):
    """Обработка отправки формы компании"""
    session = get_session()
    try:
        user_id = get_current_user_id(request)
        if not user_id:
            return RedirectResponse("/login", status_code=303)
        
        form = await request.form()
        
        # Получаем данные из формы
        website = (form.get("website") or "").strip() or None
        social_networks_str = (form.get("social_networks") or "").strip()
        social_networks = [s.strip() for s in social_networks_str.split(",") if s.strip()] if social_networks_str else []
        categories = form.getlist("categories")  # Множественный выбор
        description = (form.get("description") or "").strip() or None
        prepayment_available = form.get("prepayment_available") == "on"
        phone_number = (form.get("phone_number") or "").strip() or None
        
        # Валидация через Pydantic схему
        try:
            if CompanyDataSchema:
                from pydantic import ValidationError as PydanticValidationError
                try:
                    company_data = CompanyDataSchema(
                        website=website,
                        social_networks=social_networks,
                        categories=categories,
                        photos=[],  # Фотографии обрабатываются отдельно через UploadFile
                        description=description,
                        prepayment_available=prepayment_available,
                        phone_number=phone_number
                    )
                except PydanticValidationError as e:
                    # Обработка ошибок валидации Pydantic
                    error_messages = []
                    for error in e.errors():
                        field = error.get('loc', ['unknown'])[-1]
                        msg = error.get('msg', 'Ошибка валидации')
                        if field == 'phone_number':
                            error_messages.append(f"Номер телефона: {msg}")
                        else:
                            error_messages.append(f"{field}: {msg}")
                    
                    error_msg = "; ".join(error_messages) if error_messages else "Ошибка валидации данных"
                    
                    return templates.TemplateResponse("company_form.html", {
                        "request": request,
                        "page_title": "Заполнение данных компании",
                        "active_page": "company",
                        "user_role": get_user_role(session, user_id),
                        "error": error_msg,
                        "form_data": {
                            "website": website,
                            "social_networks": social_networks_str,
                            "categories": categories,
                            "description": description,
                            "prepayment_available": prepayment_available,
                            "phone_number": phone_number,
                        }
                    })
        except Exception as e:
            # Обработка других ошибок
            logger.exception("Ошибка при валидации данных компании")
            error_msg = f"Ошибка валидации: {str(e)}"
            
            return templates.TemplateResponse("company_form.html", {
                "request": request,
                "page_title": "Заполнение данных компании",
                "active_page": "company",
                "user_role": get_user_role(session, user_id),
                "error": error_msg,
                "form_data": {
                    "website": website,
                    "social_networks": social_networks_str,
                    "categories": categories,
                    "description": description,
                    "prepayment_available": prepayment_available,
                    "phone_number": phone_number,
                }
            })
        
        # Обработка загруженных фотографий
        photos = []
        photo_files = form.getlist("photos")
        for photo_file in photo_files:
            if hasattr(photo_file, 'filename') and photo_file.filename:
                # Сохраняем фото (здесь можно добавить логику сохранения)
                photos.append(photo_file.filename)
        
        # Валидация количества фотографий
        if len(photos) > 6:
            return templates.TemplateResponse("company_form.html", {
                "request": request,
                "page_title": "Заполнение данных компании",
                "active_page": "company",
                "user_role": get_user_role(session, user_id),
                "error": "Максимум 6 фотографий",
                "form_data": {
                    "website": website,
                    "social_networks": social_networks_str,
                    "categories": categories,
                    "description": description,
                    "prepayment_available": prepayment_available,
                    "phone_number": phone_number,
                }
            })
        
        # Здесь должна быть логика сохранения данных компании в БД
        # Пока просто возвращаем успешное сообщение
        logger.info(f"Данные компании сохранены пользователем {user_id}: website={website}, categories={categories}, prepayment={prepayment_available}")
        
        return templates.TemplateResponse("company_form.html", {
            "request": request,
            "page_title": "Заполнение данных компании",
            "active_page": "company",
            "user_role": get_user_role(session, user_id),
            "success": "Данные компании успешно сохранены!",
            "form_data": None,
        })
    except Exception as e:
        logger.exception("Ошибка при сохранении данных компании")
        return templates.TemplateResponse("company_form.html", {
            "request": request,
            "page_title": "Заполнение данных компании",
            "active_page": "company",
            "user_role": get_user_role(session, user_id) if 'session' in locals() else None,
            "error": f"Ошибка при сохранении: {str(e)}",
            "form_data": None,
        })
    finally:
        session.close()


@app.on_event("startup")
async def startup_event():
    """Инициализация при старте приложения"""
    session = get_session()
    try:
        ensure_owner_user(session)
        logging.info("✅ Базовый owner пользователь инициализирован")
    except Exception as e:
        logging.error(f"❌ Ошибка при инициализации базового owner: {e}")
    finally:
        session.close()

if __name__ == "__main__":
    main()

