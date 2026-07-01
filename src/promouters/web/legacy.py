"""Compatibility routes for paths carried over from PythonProject2 admin."""
from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, Form, Query
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from promouters.db.session import get_db
from promouters.models.service_ops import Order
from promouters.models.users import Role, User
from promouters.services import cash as cash_svc
from promouters.services import orders as orders_svc
from promouters.services import statistics as stats_svc
from promouters.web.deps import clear_auth_cookies, require_roles


router = APIRouter()


def _parse_optional_int(value: str | int | None) -> int | None:
    """Parse a query-string value to int; returns None for empty/invalid strings."""
    if value is None or isinstance(value, int):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _xlsx_response(workbook: Workbook, filename: str) -> StreamingResponse:
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _styled_sheet(workbook: Workbook, title: str, headers: list[str]):
    sheet = workbook.active if workbook.active.title == "Sheet" and workbook.active.max_row == 1 else workbook.create_sheet()
    sheet.title = title
    sheet.append(headers)
    fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    return sheet


def _find_master_by_tg(db: Session, master_tg_id: int) -> User | None:
    return db.scalar(
        select(User)
        .join(Role)
        .where(User.tg_id == master_tg_id, Role.code == "master")
    )


@router.get("/logout")
async def legacy_logout():
    response = RedirectResponse("/admin/login", status_code=302)
    clear_auth_cookies(response)
    return response


@router.get("/create-order")
async def legacy_create_order_redirect():
    return RedirectResponse("/admin/orders/create", status_code=302)


@router.get("/order-search")
async def legacy_order_search_redirect(phone: str | None = None):
    target = "/admin/orders/search"
    if phone:
        target = f"{target}?phone={phone}"
    return RedirectResponse(target, status_code=302)


@router.get("/branch-orders")
async def legacy_branch_orders_redirect():
    return RedirectResponse("/admin/orders", status_code=302)


@router.get("/branch-sd")
async def legacy_branch_sd_redirect():
    return RedirectResponse("/admin/sd", status_code=302)


@router.get("/branch-cash")
async def legacy_branch_cash_redirect():
    return RedirectResponse("/admin/cash", status_code=302)


@router.get("/branch-masters")
async def legacy_branch_masters_redirect():
    return RedirectResponse("/admin/masters", status_code=302)


@router.get("/branch-stats")
async def legacy_branch_stats_redirect():
    return RedirectResponse("/admin/stats", status_code=302)


@router.get("/order/{order_id}")
async def legacy_order_redirect(order_id: int):
    return RedirectResponse(f"/admin/orders/{order_id}", status_code=302)


@router.get("/order/{order_id}/bso-file")
async def legacy_order_bso_file(order_id: int, db: Session = Depends(get_db)):
    order = orders_svc.get_order(db, order_id)
    if order and order.bso_file_path:
        return RedirectResponse(f"/media/{order.bso_file_path}", status_code=302)
    return RedirectResponse(f"/admin/orders/{order_id}?error=bso-missing", status_code=302)


@router.get("/order/{order_id}/receipt-file")
async def legacy_order_receipt_file(order_id: int, db: Session = Depends(get_db)):
    order = orders_svc.get_order(db, order_id)
    if order and order.receipt_file_path:
        return RedirectResponse(f"/media/{order.receipt_file_path}", status_code=302)
    return RedirectResponse(f"/admin/orders/{order_id}?error=receipt-missing", status_code=302)


@router.post("/order/{order_id}/assign")
async def legacy_order_assign(
    order_id: int,
    master_id: str | None = Form(None),
    assigned_to: str | None = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("owner", "director", "dispatcher", "branch_manager", "ad_director")),
):
    order = orders_svc.get_order(db, order_id)
    if order is None:
        return RedirectResponse("/admin/orders?error=notfound", status_code=302)
    raw_value = (master_id or assigned_to or "").strip()
    payload = {"assigned_to": int(raw_value)} if raw_value else {"assigned_to": None}
    old_assigned_to = order.assigned_to
    orders_svc.update_order(db, order, payload)

    # Notify new master via VK
    if order.assigned_to and order.assigned_to != old_assigned_to:
        from promouters.services.vk_notify import notify_master_new_order
        notify_master_new_order(order)

    return RedirectResponse(f"/admin/orders/{order_id}?updated=1", status_code=302)


@router.post("/cash/accept-order/{order_id}")
async def legacy_accept_order_cash(
    order_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("owner", "director")),
):
    pending_orders = {order.id: order for order in cash_svc.list_pending_cash_orders(db)}
    order = pending_orders.get(order_id)
    if order is not None:
        cash_svc.accept_order_cash(db, order)
    return RedirectResponse("/admin/cash?accepted=order", status_code=302)


@router.post("/cash/accept/{master_id}")
async def legacy_accept_cash_for_master(
    master_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("owner", "director")),
):
    cash_svc.accept_cash_for_master(db, master_tg_id=master_id)
    return RedirectResponse("/admin/cash?accepted=master", status_code=302)


@router.get("/master/{master_tg_id}")
async def legacy_master_redirect(master_tg_id: int, db: Session = Depends(get_db)):
    master = _find_master_by_tg(db, master_tg_id)
    if master is None:
        return RedirectResponse("/admin/masters?error=notfound", status_code=302)
    return RedirectResponse(f"/admin/masters/{master.id}", status_code=302)


@router.get("/master/{master_tg_id}/passport-photo")
async def legacy_master_passport_photo(master_tg_id: int, db: Session = Depends(get_db)):
    master = _find_master_by_tg(db, master_tg_id)
    if master and master.passport_photo_path:
        return RedirectResponse(f"/media/{master.passport_photo_path}", status_code=302)
    return RedirectResponse("/admin/masters?error=passport-missing", status_code=302)


@router.get("/export/masters")
async def export_masters(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("owner", "director")),
):
    masters = list(
        db.scalars(
            select(User)
            .join(Role)
            .where(Role.code == "master")
            .order_by(User.last_name, User.first_name)
        )
    )
    rows = [
        [
            master.tg_id or "",
            master.full_name or master.name or f"{master.first_name} {master.last_name}".strip(),
            master.branch.name if master.branch else "",
            master.phone or "",
            master.master_percentage if master.master_percentage is not None else "по сетке",
        ]
        for master in masters
    ]
    headers = ["VK/Telegram ID", "Имя", "Филиал", "Телефон", "Индивидуальный %"]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="masters.csv"'},
        )

    workbook = Workbook()
    sheet = _styled_sheet(workbook, "Мастера", headers)
    for row in rows:
        sheet.append(row)
    return _xlsx_response(workbook, "masters.xlsx")


@router.get("/export/orders")
async def export_orders(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("owner", "director")),
):
    start_dt = datetime.now() - timedelta(days=days)
    orders = orders_svc.list_orders(db, date_from=start_dt, limit=5000)
    master_ids = {order.assigned_to for order in orders if order.assigned_to}
    masters_map: dict[int, User] = {}
    if master_ids:
        masters_map = {
            master.tg_id: master
            for master in db.scalars(select(User).where(User.tg_id.in_(master_ids)))
        }
    rows = []
    for order in orders:
        master = masters_map.get(order.assigned_to) if order.assigned_to else None
        rows.append(
            [
                order.order_number or order.id,
                order.city_rel.name if order.city_rel else "",
                ", ".join(part for part in (order.street, order.house, order.flat) if part),
                order.equip_type or "",
                order.status or "",
                (
                    master.full_name
                    or master.name
                    or f"{master.first_name} {master.last_name}".strip()
                ) if master else "",
                order.sum_amount or 0,
                order.created_at.strftime("%d.%m.%Y %H:%M") if order.created_at else "",
            ]
        )
    headers = ["№", "Город", "Адрес", "Техника", "Статус", "Мастер", "Сумма", "Дата создания"]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="orders_{days}days.csv"'},
        )

    workbook = Workbook()
    sheet = _styled_sheet(workbook, "Заявки", headers)
    for row in rows:
        sheet.append(row)
    return _xlsx_response(workbook, f"orders_{days}days.xlsx")


@router.get("/stats/export")
async def export_stats(
    city_id: str | None = None,
    master_id: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    period: str = "month",
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("owner", "director", "dispatcher", "branch_manager", "ad_director")),
):
    resolved_city_id = _parse_optional_int(city_id)
    resolved_master_id = _parse_optional_int(master_id)
    if date_from and date_to:
        start_dt = datetime.fromisoformat(date_from).replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59, microsecond=999999)
    else:
        start_dt, end_dt = stats_svc.get_period_bounds(period if period in stats_svc.PERIOD_PRESETS else "month")

    dashboard = stats_svc.calculate_dashboard_stats(
        db,
        date_from=start_dt,
        date_to=end_dt,
        city_id=resolved_city_id,
        master_id=resolved_master_id,
    )

    workbook = Workbook()
    summary = _styled_sheet(workbook, "Сводка", ["Показатель", "Значение"])
    for label, value in [
        ("Заявок всего", dashboard["cards"]["total"]),
        ("Завершено", dashboard["cards"]["completed"]),
        ("Оборот нетто", dashboard["cards"]["net_sum"]),
        ("Доля компании", dashboard["cards"]["company_share"]),
        ("Средний чек", dashboard["cards"]["avg_check"]),
        ("Конверсия", dashboard["cards"]["conversion"]),
    ]:
        summary.append([label, value])

    masters_sheet = workbook.create_sheet("Мастера")
    masters_sheet.append(["Мастер", "Заявок", "Нетто", "Средний чек", "Конверсия"])
    for item in dashboard["masters"]:
        masters_sheet.append([item["name"], item["count"], item["net_sum"], item["avg_check"], item["conversion"]])

    cities_sheet = workbook.create_sheet("Города")
    cities_sheet.append(["Город", "Заявок", "Нетто"])
    for item in dashboard["cities"]:
        cities_sheet.append([item["name"], item["count"], item["net_sum"]])

    equipment_sheet = workbook.create_sheet("Направления")
    equipment_sheet.append(["Направление", "Заявок", "Нетто", "Средний чек", "Отказов"])
    for item in dashboard["equipment"]:
        equipment_sheet.append([item["name"], item["count"], item["net_sum"], item["avg_check"], item["refused"]])

    daily_sheet = workbook.create_sheet("По дням")
    daily_sheet.append(["Дата", "Завершено", "Нетто"])
    for item in dashboard["daily"]:
        daily_sheet.append([item["date"], item["count"], item["net_sum"]])

    return _xlsx_response(workbook, "stats.xlsx")


@router.get("/api/stats/extended")
async def extended_stats_json(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("owner", "director", "dispatcher", "branch_manager", "ad_director")),
):
    start_dt, end_dt = stats_svc.get_period_bounds("last_30")
    dashboard = stats_svc.calculate_dashboard_stats(db, date_from=start_dt, date_to=end_dt)
    payload = {
        "equipment": [
            {"type": item["name"], "count": item["count"], "total_sum": item["net_sum"]}
            for item in dashboard["equipment"]
        ],
        "cities": [
            {"city": item["name"], "count": item["count"], "total_sum": item["net_sum"]}
            for item in dashboard["cities"]
        ],
        "revenue": [
            {"date": item["date"], "revenue": item["net_sum"]}
            for item in dashboard["daily"]
        ],
        "avg_check": dashboard["cards"]["avg_check"],
        "conversion": dashboard["cards"]["conversion"],
        "company_share": dashboard["cards"]["company_share"],
    }
    return JSONResponse(payload)
